"""
Simplified xlwt interface
"""
from datetime import date, datetime
from io import BytesIO

from django.utils import timezone
from openpyxl import Workbook, load_workbook, styles
from openpyxl.cell import Cell
from openpyxl.styles.fonts import DEFAULT_FONT
from openpyxl.worksheet.worksheet import Worksheet
from pytz import UTC


def empty_workbook():
    wkbk: Workbook = Workbook()
    wkbk.remove(wkbk.worksheets[0])
    return wkbk


font_size = 12
default_font = styles.Font(size=font_size, name=DEFAULT_FONT.name)
header_font = styles.Font(bold=True, size=font_size, name=DEFAULT_FONT.name)
url_font = styles.Font(color=styles.colors.BLUE, size=font_size, name=DEFAULT_FONT.name)


def add_sheet_with_header_row(wkbk: Workbook, name: str, headers: list[str], contents: list[list[str]]):
    """
    Utility function for adding sheet to xlwt workbook.
    """
    wksh: Worksheet = wkbk.create_sheet(title=name)

    border = styles.Border(
        left=styles.Side(border_style="thin"),
        right=styles.Side(border_style="thin"),
        top=styles.Side(border_style="thin"),
        bottom=styles.Side(border_style="thin"),
    )

    alignment = styles.Alignment(vertical="center")
    wrapped_alignment = styles.Alignment(vertical="center", wrapText=True)
    date_format = "YYYY/MM/DD"

    for c_idx, header in enumerate(headers, start=1):
        cell: Cell = wksh.cell(row=1, column=c_idx, value=header)
        cell.font = header_font
        cell.border = border

    header_row_count = 1

    for r_idx, row in enumerate(contents, start=1 + header_row_count):
        normal_row_height = font_size
        row_height = normal_row_height
        for c_idx, val in enumerate(row, start=1):
            cell: Cell = wksh.cell(row=r_idx, column=c_idx)
            cell.border = border
            cell.alignment = alignment
            cell.font = default_font

            if isinstance(val, str):
                # normalise newlines to style expected by Excel
                val = val.replace("\r\n", "\n")

            if isinstance(val, datetime | date):
                cell.number_format = date_format
                if isinstance(val, datetime):
                    if timezone.is_aware(val):
                        val = timezone.make_naive(val, UTC)
            else:
                if isinstance(val, str) and "\n" in val:
                    # This is needed or Excel displays box character for newlines.
                    cell.alignment = wrapped_alignment
                    # Set height to be able to see all lines
                    row_height = max(row_height, font_size * (val.count("\n") + 1))
                if looks_like_url(val):
                    val = f'=HYPERLINK("{val}"; "{val}")'
                    cell.font = url_font
            cell.value = val
        if row_height > normal_row_height:
            wksh.row_dimensions[r_idx].height = row_height


def looks_like_url(val):
    return (
        isinstance(val, str)
        and " " not in val
        and "\n" not in val
        and (val.startswith("http://") or val.startswith("https://"))
    )


def workbook_to_bytes(wkbk: Workbook) -> bytes:
    s = BytesIO()
    wkbk.save(s)
    s.seek(0)
    return s.read()


def workbook_from_bytes(content: bytes) -> Workbook:
    s = BytesIO(content)
    return load_workbook(s)

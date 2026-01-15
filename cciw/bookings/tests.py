from __future__ import annotations

import io
import json
from datetime import date, datetime, timedelta
from decimal import Decimal
from typing import Literal, assert_never
from unittest import mock

import openpyxl
import pytest
import vcr
from django.conf import settings
from django.core import mail, signing
from django.db import models
from django.test.client import Client
from django.test.utils import override_settings
from django.urls import reverse
from django.utils import timezone
from django_functest import FuncBaseMixin, Upload
from freezegun import freeze_time
from hypothesis import example, given
from hypothesis import strategies as st

from cciw.bookings.email import EmailVerifyTokenGenerator, VerifyExpired, VerifyFailed, send_payment_reminder_emails
from cciw.bookings.hooks import paypal_payment_received, unrecognised_payment
from cciw.bookings.mailchimp import get_status
from cciw.bookings.middleware import BOOKING_COOKIE_SALT
from cciw.bookings.models import (
    AccountTransferPayment,
    ApprovalNeededType,
    Booking,
    BookingAccount,
    BookingState,
    ManualPayment,
    ManualPaymentType,
    Payment,
    PaymentSource,
    Price,
    PriceType,
    RefundPayment,
    add_basket_to_queue,
    build_paypal_custom_field,
)
from cciw.bookings.models.constants import Sex
from cciw.bookings.models.expiry import expire_bookings
from cciw.bookings.models.prices import are_prices_set_for_year
from cciw.bookings.models.problems import ApprovalStatus, BookingApproval, get_booking_problems
from cciw.bookings.models.queue import (
    BookingQueueEntry,
    QueueEntryActionLogType,
    add_queue_cutoffs,
    allocate_bookings_now,
    allocate_places_and_notify,
    get_booking_queue_problems,
    get_camp_booking_queue_ranking_result,
    rank_queue_bookings,
)
from cciw.bookings.models.yearconfig import YearConfig, YearConfigFetcher, get_booking_open_data
from cciw.bookings.utils import camp_bookings_to_spreadsheet, payments_to_spreadsheet
from cciw.cciwmain.models import Camp
from cciw.cciwmain.tests import factories as camps_factories
from cciw.cciwmain.tests.mailhelpers import path_and_query_to_url, read_email_url
from cciw.officers.tests import factories as officers_factories
from cciw.sitecontent.models import HtmlChunk
from cciw.utils.tests.base import disable_logging
from cciw.utils.tests.db import refresh
from cciw.utils.tests.factories import Auto
from cciw.utils.tests.webtest import SeleniumBase, WebTestBase

from . import factories

ANT = ApprovalNeededType


class IpnMock:
    payment_status = "Completed"
    business = settings.PAYPAL_RECEIVER_EMAIL


# == Mixins to reduce duplication ==


class BookingLogInMixin:
    booker_email = "booker@bookers.com"

    def booking_login(self, add_account_details=True, shortcut=None) -> BookingAccount:
        account: BookingAccount | None = getattr(self, "_logged_in_account", None)
        if account is not None:
            return account

        if shortcut is None:
            shortcut = self.is_full_browser_test

        if shortcut:
            account, _ = BookingAccount.objects.get_or_create(email=self.booker_email)
            self._set_signed_cookie("bookingaccount", account.id, salt=BOOKING_COOKIE_SALT)
        else:
            # Easiest way is to simulate what the user actually has to do
            self.get_url("cciw-bookings-start")
            self.fill_by_name({"email": self.booker_email})
            self.submit("[type=submit]")
            url, path, querydata = read_email_url(mail.outbox.pop(), "https://.*/booking/v/.*")
            self.get_literal_url(path_and_query_to_url(path, querydata))
            account = BookingAccount.objects.get(email=self.booker_email)

        if add_account_details:
            account.name = "Joe"
            account.address_line1 = "456 My Street"
            account.address_city = "Metrocity"
            account.address_country = "GB"
            account.address_post_code = "XYZ"
            account.phone_number = "0123 456789"
            account.save()
        self._logged_in_account = account
        return account

    def _set_signed_cookie(self, key, value, salt=""):
        value = signing.get_cookie_signer(salt=key + salt).sign(value)
        if self.is_full_browser_test:
            if not self._have_visited_page():
                self.get_url("django_functest.emptypage")
            return self._driver.add_cookie({"name": key, "value": value, "path": "/"})
        else:
            return self.app.set_cookie(key, value)


class CreateBookingWebMixin(BookingLogInMixin):
    """
    Mixin to be used for functional testing of creating bookings online. It
    creates `self.camp` and `self.camp_2` and provides other utility methods.
    """

    # For other, model level tests, we prefer explicit use of factories
    # for the things under test.

    def setUp(self):
        super().setUp()
        # We have to control freeze_time usage here for this mixin,
        # to avoid some issues if the monkey patching runs too early.
        # (specifically:
        #   fontTools/misc/loggingTools.py:292: in __init__
        #    TypeError: fake_perf_counter() takes 0 positional arguments but 1 was given)

        import cciw.officers.views  # noqa - trigger some imports

        # For most tests it is easier to assume a fixed date for booking,
        # before the summer. Camp dates are then created relative to that.
        freezer = freeze_time("2026-02-01")
        freezer.start()
        self.freezer = freezer
        self.today = date.today()
        self.create_camps()

    def tearDown(self):
        super().tearDown()
        self.freezer.stop()

    camp_minimum_age = 11
    camp_maximum_age = 17

    def create_camps(self):
        if hasattr(self, "camp"):
            return
        # Need to create a Camp that we can choose i.e. is in the future.
        delta_days = 30
        start_date = self.today + timedelta(delta_days)
        self.camp = camps_factories.create_camp(
            camp_name="Blue",
            minimum_age=self.camp_minimum_age,
            maximum_age=self.camp_maximum_age,
            start_date=start_date,
        )
        self.camp_2 = camps_factories.create_camp(
            camp_name="Red",
            minimum_age=self.camp_minimum_age,
            maximum_age=self.camp_maximum_age,
            start_date=start_date + timedelta(days=7),
        )

    def open_bookings(self, *, for_data_entry_only: bool = False) -> None:
        self.add_prices()
        if for_data_entry_only:
            self.create_year_config(open_for_booking=False, open_for_data_entry=True)
        else:
            self.create_year_config(open_for_booking=True)

    def ensure_bookings_open(self, *, year: int | None = None):
        year = year or self.camp.year
        assert isinstance(year, int)
        if not are_prices_set_for_year(year):
            self.add_prices()
        if not YearConfig.objects.filter(year=year).exists():
            self.create_year_config(year=year, open_for_booking=True, open_for_data_entry=True)
        assert get_booking_open_data(year).is_open_for_booking

    def create_year_config(
        self, *, open_for_data_entry: bool = True, open_for_booking: bool = True, year: int | None = None
    ) -> YearConfig:
        time_for_booking = "past" if open_for_booking else "future"
        date_for_data_entry = "past" if open_for_data_entry else "future"
        year: int = year or self.camp.year
        return factories.create_year_config(
            year=year, bookings_open_for_booking_on=time_for_booking, bookings_open_for_entry_on=date_for_data_entry
        )

    def add_prices(self):
        if hasattr(self, "price_full"):
            return
        year = self.camp.year
        (
            self.price_full,
            self.price_2nd_child,
            self.price_3rd_child,
            self.price_booking_fee,
        ) = factories.create_prices(year=year)

    def create_booking(
        self,
        *,
        shortcut: bool = Auto,
        camp: Camp = Auto,
        first_name: str = Auto,
        last_name: str = Auto,
        name: str = Auto,
        sex="m",
        birth_date: date = Auto,
        serious_illness: bool = False,
        price_type=PriceType.FULL,
    ) -> Booking:
        """
        Creates a booking, normally using views.
        """
        if camp is Auto:
            camp = self.camp
        if shortcut is Auto:
            # To speed up full browser test, we create booking using the shortcut
            shortcut = self.is_full_browser_test

        # DWIM - we always want prices to exist and bookings to be open if we
        # call 'create_booking()':
        self.ensure_bookings_open(year=camp.year)

        data = self.get_place_details(
            camp=camp,
            first_name=first_name,
            last_name=last_name,
            name=name,
            sex=sex,
            birth_date=birth_date,
            price_type=price_type,
            serious_illness=serious_illness,
        )
        if shortcut:
            data.update(
                {
                    "account": BookingAccount.objects.get(email=self.booker_email),
                    "state": BookingState.INFO_COMPLETE,
                }
            )
            return factories.create_booking(**data)

        # Normally we use public views to create place, to ensure that they
        # are created in the same way that a user would.
        old_booking_ids = list(Booking.objects.values_list("id", flat=True))

        self.get_url("cciw-bookings-add_place")
        # Sanity check:
        self.assertTextPresent("Please enter the details needed to book a place on a camp")
        self.fill_by_name(data)
        self.submit("#id_save_btn")
        self.assertUrlsEqual(reverse("cciw-bookings-basket_list_bookings"))
        new_booking = Booking.objects.exclude(id__in=old_booking_ids).get()
        return new_booking

    def get_place_details(
        self,
        *,
        first_name: str = Auto,
        last_name: str = Auto,
        name: str = Auto,
        camp: Camp = Auto,
        birth_date: date = Auto,
        serious_illness: bool = False,
        price_type=PriceType.FULL,
        sex="m",
    ) -> dict:
        if name is not Auto:
            assert first_name is Auto
            assert last_name is Auto
            first_name, last_name = name.split(" ")
        else:
            first_name = first_name or "Frédéric"
            last_name = last_name or "Bloggs"
        if camp is Auto:
            camp = self.camp
        if birth_date is Auto:
            birth_date = date(camp.year - 14, 1, 1)
        if sex is Auto:
            sex = "m"
        return {
            # Order follows order in form.
            "camp": camp,
            "price_type": price_type,
            "first_name": first_name,
            "last_name": last_name,
            "sex": sex,
            "birth_date": birth_date,
            "address_line1": "123 My street",
            "address_city": "Metrocity",
            "address_country": "GB",
            "address_post_code": "ABC 123",
            "contact_name": "Mr Father",
            "contact_line1": "98 Main Street",
            "contact_city": "Metrocity",
            "contact_country": "GB",
            "contact_post_code": "ABC 456",
            "contact_phone_number": "01982 987654",
            "gp_name": "Doctor Who",
            "gp_line1": "The Tardis",
            "gp_city": "London",
            "gp_country": "GB",
            "gp_post_code": "SW1 1PQ",
            "gp_phone_number": "01234 456789",
            "medical_card_number": "asdfasdf",
            "last_tetanus_injection_date": date(camp.year - 5, 2, 3),
            "serious_illness": serious_illness,
            "agreement": True,
        }

    def fill(self, data, scroll=True):
        # Accept more things than super().fill()
        # This allows us to write `get_place_details()` above in a way
        # that means that data can be easily passed to Factory.create_booking()
        data2 = {}
        for k, v in data.items():
            if isinstance(v, models.Model):
                # Allow using Camp instances
                data2[k] = v.id
            elif isinstance(v, date):
                data2[k] = v.isoformat()
            else:
                data2[k] = v
        return super().fill(data2, scroll=scroll)


class BookingBaseMixin:
    # Constants used in 'assertTextPresent' and 'assertTextAbsent', the latter
    # being prone to false positives if a constant isn't used.
    ABOVE_MAXIMUM_AGE = "above the maximum age"
    BELOW_MINIMUM_AGE = "below the minimum age"
    CAMP_CLOSED_FOR_BOOKINGS = "This camp is closed for bookings"
    CANNOT_USE_2ND_CHILD = "You cannot use a 2nd child discount"
    CANNOT_USE_MULTIPLE_DISCOUNT_FOR_ONE_CAMPER = "only one place may use a 2nd/3rd child discount"
    MULTIPLE_2ND_CHILD_WARNING = "You have multiple places at '2nd child"
    MULTIPLE_FULL_PRICE_WARNING = "You have multiple places at 'Full price"
    NOT_ENOUGH_PLACES = "There are not enough places left on this camp"
    NOT_ENOUGH_PLACES_FOR_BOYS = "There are not enough places for boys left on this camp"
    NOT_ENOUGH_PLACES_FOR_GIRLS = "There are not enough places for girls left on this camp"
    NO_PLACES_LEFT = "There are no places left on this camp"
    NO_PLACES_LEFT_FOR_BOYS = "There are no places left for boys"
    NO_PLACES_LEFT_FOR_GIRLS = "There are no places left for girls"
    BOOKING_IS_NOT_OPEN = "Booking is not yet open"
    LAST_TETANUS_INJECTION_DATE_REQUIRED = "last tetanus injection"
    BOOKINGS_WILL_EXPIRE = "you have 24 hours to complete payment online"

    def setUp(self):
        super().setUp()
        HtmlChunk.objects.get_or_create(name="bookingform_post_to", menu_link=None)
        HtmlChunk.objects.get_or_create(name="booking_secretary_address", menu_link=None)


# == Test cases ==

# Most tests are against views, instead of model-based tests.
# get_booking_problems(), for instance, is tested especially in
# TestListBookings. In theory this could be tested using model-based tests
# instead, but the way that multiple bookings and the basket/shelf interact mean
# we need to test the view code as well. It would probably be good to rewrite
# using a class like "CheckoutPage", which combines shelf and basket bookings,
# and some of the logic in the list bookings view. There is also the advantage that
# using self.create_booking() (which uses a view) ensures Booking instances are
# created the same way a user would.


@pytest.mark.django_db
def test_Camp_open_for_bookings():
    today = date.today()
    camp = camps_factories.create_camp(start_date=today + timedelta(days=10))
    assert camp.open_for_bookings(today)
    assert camp.open_for_bookings(camp.start_date)
    assert not camp.open_for_bookings(camp.start_date + timedelta(days=1))

    camp.last_booking_date = today
    assert camp.open_for_bookings(today)
    assert not camp.open_for_bookings(today + timedelta(days=1))


@pytest.mark.django_db
def test_BookingAccount_balance_due(django_assert_num_queries):
    year_config = create_year_config_for_queue_tests()
    year: int = year_config.year
    factories.create_prices(year=year, full_price=100)
    camp = camps_factories.create_camp(year=year)
    config_fetcher = YearConfigFetcher()
    config_fetcher.lookup_year(year)  # prefetch to avoid complicating assertNumQueries below

    def assert_account_balance(expected: Decimal | int, *, full: bool = False):
        expected = Decimal(expected)

        if full:
            today = None
        else:
            today = date.today()
        for use_prefetch_related_for_get_account in [True, False]:
            if use_prefetch_related_for_get_account:
                # Tests that the other code paths in get_balance/BookingManager.payable
                # work.
                account = BookingAccount.objects.filter(id=booking_account_id).prefetch_related("bookings")[0]
            else:
                account = BookingAccount.objects.get(id=booking_account_id)
            with django_assert_num_queries(num=0 if use_prefetch_related_for_get_account else 2):
                assert account.get_balance(today=today, config_fetcher=config_fetcher) == expected
                assert account.get_balance(today=today, config_fetcher=config_fetcher) == expected

    # Data entry
    with freeze_time(year_config.bookings_open_for_entry_on + timedelta(days=1)):
        booking = factories.create_booking(camp=camp)
        booking_account_id: int = booking.account.id
        assert_account_balance(0)

    # "Book" button
    with freeze_time(year_config.bookings_open_for_booking_on):
        booking.add_to_queue(by_user=booking.account)
        assert_account_balance(0)

    # Confirmed by booking secretary
    with freeze_time(year_config.bookings_initial_notifications_on):
        allocate_bookings_now([booking])
        assert_account_balance(0)

        # Place should be booked
        booking.refresh_from_db()
        assert booking.state == BookingState.BOOKED

    # Before full payment due:
    with freeze_time(year_config.payments_due_on - timedelta(days=1)):
        # balance should be zero
        assert_account_balance(0)

        # But for full amount, they still owe 100 (full price)
        assert_account_balance(100, full=True)

        # Test some model methods:
        assert len(booking.account.bookings.payable()) == 1

    with freeze_time(year_config.payments_due_on):
        assert_account_balance(100)
        assert_account_balance(100, full=True)


class BookingIndexBase(BookingBaseMixin, FuncBaseMixin):
    def test_show_with_no_prices(self):
        camp = camps_factories.create_camp()
        year = camp.year
        self.get_url("cciw-bookings-index")
        self.assertTextPresent(f"Prices for {year} have not been finalised yet")

    def test_show_with_prices(self):
        camp = camps_factories.create_camp()
        year = camp.year
        factories.create_prices(year=year, full_price=100)
        self.get_url("cciw-bookings-index")
        self.assertTextPresent("£100")

        # No YearConfig has been set, so we can't show much "booking process info"
        self.assertTextPresent(f"The dates for the booking process for {year} have not been set.")

    def test_show_with_yearconfig(self):
        camp = camps_factories.create_camp()
        year = camp.year
        config = factories.create_year_config(year=year)
        assert config.bookings_open_for_booking_on is not None

        self.get_url("cciw-bookings-index")
        # YearConfig has been set, so we show full "booking process info":
        self.assertTextPresent("To ensure a fair and reliable booking process")


class TestBookingIndexWT(BookingIndexBase, WebTestBase):
    pass


class BookingStartBase(BookingBaseMixin, CreateBookingWebMixin, FuncBaseMixin):
    urlname = "cciw-bookings-start"

    def submit(self, css_selector="[type=submit]"):
        return super().submit(css_selector)

    def test_show_form(self):
        self.get_url(self.urlname)
        self.assertTextPresent("id_email")

    def test_complete_form(self):
        assert BookingAccount.objects.all().count() == 0
        self.get_url(self.urlname)
        self.fill_by_name({"email": "booker@bookers.com"})
        self.submit()
        assert BookingAccount.objects.all().count() == 0
        assert len(mail.outbox) == 1

    def test_complete_form_existing_email(self):
        BookingAccount.objects.create(email="booker@bookers.com")
        assert BookingAccount.objects.all().count() == 1
        self.get_url(self.urlname)
        self.fill_by_name({"email": "booker@bookers.com"})
        self.submit()
        assert BookingAccount.objects.all().count() == 1
        assert len(mail.outbox) == 1

    def test_complete_form_existing_email_different_case(self):
        BookingAccount.objects.create(email="booker@bookers.com")
        assert BookingAccount.objects.all().count() == 1
        self.get_url(self.urlname)
        self.fill_by_name({"email": "BOOKER@bookers.com"})
        self.submit()
        assert BookingAccount.objects.all().count() == 1
        assert len(mail.outbox) == 1

    def test_skip_if_logged_in(self):
        # This assumes verification process works
        # Check redirect to step 3 - account details
        self.booking_login(add_account_details=False)
        self.get_url(self.urlname)
        self.assertUrlsEqual(reverse("cciw-bookings-account_details"))

    def test_skip_if_account_details(self):
        # Check redirect to step 4 - add place
        self.booking_login()
        self.get_url(self.urlname)
        self.assertUrlsEqual(reverse("cciw-bookings-add_place"))

    def test_skip_if_has_place_details(self):
        # Check redirect to overview
        account = self.booking_login()
        factories.create_booking(account=account)
        self.get_url(self.urlname)
        self.assertUrlsEqual(reverse("cciw-bookings-account_overview"))


class TestBookingStartWT(BookingStartBase, WebTestBase):
    pass


class TestBookingStartSL(BookingStartBase, SeleniumBase):
    pass


class BookingVerifyBase(BookingBaseMixin, FuncBaseMixin):
    def submit(self, css_selector="[type=submit]"):
        return super().submit(css_selector)

    def _read_email_verify_email(self, email):
        return read_email_url(email, "https://.*/booking/v/.*")

    def _start(self, email="booker@bookers.com"):
        # Assumes booking_start works:
        self.get_url("cciw-bookings-start")
        self.fill_by_name({"email": email})
        self.submit()

    def test_verify_correct(self):
        """
        Test the email verification stage when the URL is correct
        """
        self._start(email="booker@bookers.com")
        url, path, querydata = self._read_email_verify_email(mail.outbox[-1])
        self.get_literal_url(path_and_query_to_url(path, querydata))
        self.assertUrlsEqual(reverse("cciw-bookings-account_details"))
        self.assertTextPresent("Logged in as booker@bookers.com! You will stay logged in for two weeks")
        account = BookingAccount.objects.get(email="booker@bookers.com")
        assert account.last_login_at is not None
        assert account.first_login_at is not None

    def _add_booking_account_address(self, email="booker@bookers.com"):
        account, _ = BookingAccount.objects.get_or_create(email=email)
        account.name = "Joe"
        account.address_line1 = "Home"
        account.address_city = "My city"
        account.address_country = "GB"
        account.address_post_code = "XY1 D45"
        account.save()

    def test_verify_correct_and_has_details(self):
        """
        Test the email verification stage when the URL is correct and the
        account already has name and address
        """
        self._start()
        self._add_booking_account_address()
        url, path, querydata = self._read_email_verify_email(mail.outbox[-1])
        self.get_literal_url(path_and_query_to_url(path, querydata))
        self.assertUrlsEqual(reverse("cciw-bookings-add_place"))

    def test_verify_correct_and_has_old_details(self):
        """
        Test the email verification stage when the URL is correct and the
        account already has name and address, but they haven't logged in
        for 'a while'.
        """
        self._start()
        self._add_booking_account_address()
        account = BookingAccount.objects.get(email="booker@bookers.com")
        account.first_login_at = timezone.now() - timedelta(30 * 7)
        account.last_login_at = account.first_login_at
        account.save()

        url, path, querydata = self._read_email_verify_email(mail.outbox[-1])
        self.get_literal_url(path_and_query_to_url(path, querydata))
        self.assertUrlsEqual(reverse("cciw-bookings-account_details"))
        self.assertTextPresent("Welcome back!")
        self.assertTextPresent("Please check and update your account details")

    def test_verify_incorrect(self):
        """
        Test the email verification stage when the URL is incorrect
        """
        self._start()

        # The following will trigger a BadSignature
        url, path, querydata = self._read_email_verify_email(mail.outbox[-1])
        querydata["bt"] = "a000" + querydata["bt"]
        self.get_literal_url(path_and_query_to_url(path, querydata))
        self.assertTextPresent("failed", within="title")

        # This will trigger a base64 decode error:
        url, path, querydata = self._read_email_verify_email(mail.outbox[-1])
        querydata["bt"] = "XXX" + querydata["bt"]
        self.get_literal_url(path_and_query_to_url(path, querydata))
        self.assertTextPresent("failed", within="title")

        # This will trigger a UnicodeDecodeError
        url, path, querydata = self._read_email_verify_email(mail.outbox[-1])
        querydata["bt"] = "xxxx"
        self.get_literal_url(path_and_query_to_url(path, querydata))
        self.assertTextPresent("failed", within="title")


class TestBookingVerifyWT(BookingVerifyBase, WebTestBase):
    pass


class TestBookingVerifySL(BookingVerifyBase, SeleniumBase):
    pass


class TestPaymentReminderEmails(BookingBaseMixin, WebTestBase):
    def _create_booking(self):
        booking = factories.create_booking()
        allocate_bookings_now([booking])
        booking: Booking = Booking.objects.get(id=booking.id)
        assert len(BookingAccount.objects.payments_due()) == 1
        return booking

    def test_payment_reminder_email(self):
        booking = self._create_booking()
        mail.outbox = []
        send_payment_reminder_emails()
        assert len(mail.outbox) == 1
        m = mail.outbox[0]
        assert "You have payments due" in m.body
        assert "[CCIW] Payment due" == m.subject
        url, path, querydata = read_email_url(m, "https://.*/booking/p.*")
        self.get_literal_url(path_and_query_to_url(path, querydata))
        self.assertUrlsEqual(reverse("cciw-bookings-pay"))
        self.assertTextPresent(booking.account.get_balance_due_now())

    def test_payment_reminder_email_link_expired(self):
        self._create_booking()
        mail.outbox = []
        send_payment_reminder_emails()
        m = mail.outbox[0]
        url, path, querydata = read_email_url(m, "https://.*/booking/p.*")

        with override_settings(BOOKING_EMAIL_VERIFY_TIMEOUT=timedelta(days=-1)):
            self.get_literal_url(path_and_query_to_url(path, querydata))

        # link expired, new email should be sent.
        self.assertUrlsEqual(reverse("cciw-bookings-link_expired_email_sent"))
        assert len(mail.outbox) == 2
        m2 = mail.outbox[1]

        url2, path2, querydata2 = read_email_url(m2, "https://.*/booking/p.*")
        self.get_literal_url(path_and_query_to_url(path2, querydata2))
        self.assertUrlsEqual(reverse("cciw-bookings-pay"))


class AccountDetailsBase(BookingBaseMixin, BookingLogInMixin, FuncBaseMixin):
    urlname = "cciw-bookings-account_details"
    submit_css_selector = "[type=submit]"

    def submit(self, css_selector=submit_css_selector):
        return super().submit(css_selector)

    def test_redirect_if_not_logged_in(self):
        self.get_url(self.urlname)
        self.assertUrlsEqual(reverse("cciw-bookings-not_logged_in"))

    def test_show_if_logged_in(self):
        self.booking_login(add_account_details=False)
        self.get_url(self.urlname)
        self.assertUrlsEqual(reverse(self.urlname))

    def test_missing_name(self):
        self.booking_login(add_account_details=False)
        self.get_url(self.urlname)
        self.submit_expecting_html5_validation_errors()
        self.assertTextPresent("This field is required")

    @mock.patch("cciw.bookings.mailchimp.update_newsletter_subscription")
    def test_complete(self, UNS_func):
        """
        Test that we can complete the account details page
        """
        account = self.booking_login(add_account_details=False)
        self.get_url(self.urlname)
        self._fill_in_account_details()
        self.submit()
        account.refresh_from_db()
        assert account.name == "Mr Booker"
        assert UNS_func.call_count == 0

    @mock.patch("cciw.bookings.mailchimp.update_newsletter_subscription")
    def test_news_letter_subscribe(self, UNS_func):
        account = self.booking_login(add_account_details=False)
        self.get_url(self.urlname)
        self._fill_in_account_details()
        self.fill({"#id_subscribe_to_newsletter": True})
        self.submit()
        account.refresh_from_db()
        assert account.subscribe_to_newsletter
        assert UNS_func.call_count == 1

    def test_subscribe_to_mailings_unselected(self):
        account = self.booking_login(add_account_details=False)
        self.get_url(self.urlname)
        #  Initial value should be NULL - we haven't asked.
        assert account.subscribe_to_mailings is None
        assert account.include_in_mailings is True
        self._fill_in_account_details()
        self.submit()
        account.refresh_from_db()
        # The form should default to 'False'. As soon as this
        # page has been submitted, we *have* asked the question
        # and they have said 'no' by not selecting the box.
        assert account.subscribe_to_mailings is False
        assert account.include_in_mailings is False

    def test_subscribe_to_mailings_selected(self):
        account = self.booking_login(add_account_details=False)
        self.get_url(self.urlname)
        self._fill_in_account_details()
        self.fill({"#id_subscribe_to_mailings": True})
        self.submit()
        account.refresh_from_db()
        assert account.subscribe_to_mailings is True
        assert account.include_in_mailings is True

    def _fill_in_account_details(self):
        self.fill_by_name(
            {
                "name": "Mr Booker",
                "address_line1": "123, A Street",
                "address_city": "Metrocity",
                "address_country": "GB",
                "address_post_code": "XY1 D45",
            }
        )

    # For updating this, see:
    # https://vcrpy.readthedocs.org/en/latest/usage.html

    @vcr.use_cassette("cciw/bookings/fixtures/vcr_cassettes/subscribe.yaml", ignore_localhost=True)
    def test_subscribe(self):
        account = self.booking_login(add_account_details=False)
        self.get_url(self.urlname)
        self._fill_in_account_details()
        self.fill_by_name({"subscribe_to_newsletter": True})
        self.submit()
        account.refresh_from_db()
        assert account.subscribe_to_newsletter
        assert get_status(account) == "subscribed"

    @vcr.use_cassette("cciw/bookings/fixtures/vcr_cassettes/unsubscribe.yaml", ignore_localhost=True)
    def test_unsubscribe(self):
        account = self.booking_login()
        account.subscribe_to_newsletter = True
        account.save()

        self.get_url(self.urlname)
        self.fill_by_name({"subscribe_to_newsletter": False})
        self.submit()
        account.refresh_from_db()
        assert not account.subscribe_to_newsletter
        assert get_status(account) == "unsubscribed"


class TestAccountDetailsWT(AccountDetailsBase, WebTestBase):
    pass


class TestAccountDetailsSL(AccountDetailsBase, SeleniumBase):
    pass


class AddPlaceBase(BookingBaseMixin, CreateBookingWebMixin, FuncBaseMixin):
    urlname = "cciw-bookings-add_place"

    SAVE_BTN = "#id_save_btn"

    submit_css_selector = SAVE_BTN

    def submit(self, css_selector=submit_css_selector):
        return super().submit(css_selector)

    def test_redirect_if_not_logged_in(self):
        self.get_url(self.urlname)
        self.assertUrlsEqual(reverse("cciw-bookings-not_logged_in"))

    def test_redirect_if_no_account_details(self):
        self.booking_login(add_account_details=False)
        self.get_url(self.urlname)
        self.assertUrlsEqual(reverse("cciw-bookings-account_details"))

    def test_show_if_logged_in(self):
        self.booking_login()
        self.get_url(self.urlname)
        self.assertUrlsEqual(reverse(self.urlname))

    def test_show_error_if_not_open(self):
        self.booking_login()
        self.get_url(self.urlname)
        self.assertTextPresent(self.BOOKING_IS_NOT_OPEN)

    def test_post_not_allowed_if_not_open(self):
        self.booking_login()
        self.get_url(self.urlname)
        assert not self.is_element_present(self.SAVE_BTN)

        self.open_bookings(for_data_entry_only=True)
        self.get_url(self.urlname)
        data = self.get_place_details()
        self.fill_by_name(data)
        # Now remove prices, just to be awkward:
        Price.objects.all().delete()
        self.submit()
        self.assertTextPresent(self.BOOKING_IS_NOT_OPEN)

    def test_allowed_if_prices_set_and_year_config_open_for_data_entry(self):
        self.booking_login()
        self.add_prices()
        self.create_year_config(open_for_data_entry=True)
        self.get_url(self.urlname)
        self.assertTextAbsent(self.BOOKING_IS_NOT_OPEN)

    def test_incomplete(self):
        self.booking_login()
        self.open_bookings()
        self.get_url(self.urlname)
        self.submit_expecting_html5_validation_errors()
        self.assertTextPresent("This field is required")

    def test_complete(self):
        account = self.booking_login()
        self.open_bookings()
        self.get_url(self.urlname)
        assert account.bookings.count() == 0
        data = self.get_place_details()
        self.fill_by_name(data)
        self.submit()
        self.assertUrlsEqual(reverse("cciw-bookings-basket_list_bookings"))

        # Did we create it?
        assert account.bookings.count() == 1

        booking = account.bookings.get()

        # Check attributes set correctly
        assert booking.amount_due == self.price_full
        assert booking.created_online
        assert not booking.publicity_photos_agreement


class TestAddPlaceWT(AddPlaceBase, WebTestBase):
    pass


class TestAddPlaceSL(AddPlaceBase, SeleniumBase):
    def _use_existing_start(self):
        self.booking_login()
        self.create_booking(shortcut=True)
        self.get_url(self.urlname)

    def assertValues(self, data):
        for k, v in data.items():
            assert self.value(k) == v

    def test_use_existing_addresses(self):
        self._use_existing_start()

        self.click(".use_previous_data")
        self.fill({"#id_copy_address_details": True, "#id_copy_contact_address_details": True})
        self.click("dialog button[name=copy]")
        self.wait_for_ajax()

        self.assertValues(
            {
                "#id_address_line1": "123 My street",
                "#id_address_country": "GB",
                "#id_address_post_code": "ABC 123",
                "#id_contact_name": "Mr Father",
                "#id_contact_line1": "98 Main Street",
                "#id_contact_country": "GB",
                "#id_contact_post_code": "ABC 456",
                "#id_first_name": "",
                "#id_gp_name": "",
                "#id_gp_line1": "",
                "#id_gp_country": "GB",
            }
        )

    def test_use_existing_gp(self):
        self._use_existing_start()

        self.click(".use_previous_data")
        self.fill({"#id_copy_gp_details": True})
        self.click("dialog button[name=copy]")
        self.wait_for_ajax()

        self.assertValues(
            {
                "#id_address_line1": "",
                "#id_address_country": "GB",
                "#id_address_post_code": "",
                "#id_contact_name": "",
                "#id_contact_line1": "",
                "#id_contact_country": "GB",
                "#id_contact_post_code": "",
                "#id_first_name": "",
                "#id_gp_name": "Doctor Who",
                "#id_gp_line1": "The Tardis",
                "#id_gp_country": "GB",
            }
        )

    def test_use_existing_camper(self):
        self._use_existing_start()

        self.click(".use_previous_data")
        self.fill({"#id_copy_camper_details": True})
        self.click("dialog button[name=copy]")
        self.wait_for_ajax()

        self.assertValues(
            {
                "#id_first_name": "Frédéric",
            }
        )

    def test_use_account_data(self):
        self._use_existing_start()

        self.click("button[name=copy_account_address_to_camper]")
        self.wait_for_ajax()
        self.assertValues(
            {
                "#id_address_line1": "456 My Street",
                "#id_address_city": "Metrocity",
                "#id_address_country": "GB",
                "#id_phone_number": "0123 456789",
                "#id_address_post_code": "XYZ",
            }
        )

        self.click("button[name=copy_account_address_to_contact_details]")
        self.wait_for_ajax()
        self.assertValues(
            {
                "#id_contact_line1": "456 My Street",
                "#id_contact_name": "Joe",
                "#id_contact_city": "Metrocity",
                "#id_contact_country": "GB",
                "#id_contact_phone_number": "0123 456789",
                "#id_contact_post_code": "XYZ",
            }
        )


class EditPlaceBase(BookingBaseMixin, CreateBookingWebMixin, FuncBaseMixin):
    # Most functionality is shared with the 'add' form, so doesn't need testing separately.

    submit_css_selector = "#id_save_btn"

    def edit_place(self, booking, expect_code=None):
        url = reverse("cciw-bookings-edit_place", kwargs={"booking_id": str(booking.id)})
        expect_errors = expect_code is not None and str(expect_code).startswith("4")
        action = lambda: self.get_literal_url(url, expect_errors=expect_errors)
        if expect_errors:
            with disable_logging():  # suppress django.request warning
                action()
        else:
            action()
        if expect_code is not None:
            self.assertCode(expect_code)

    def submit(self, css_selector=submit_css_selector):
        return super().submit(css_selector)

    def test_redirect_if_not_logged_in(self):
        self.get_url("cciw-bookings-edit_place", booking_id="1")
        self.assertUrlsEqual(reverse("cciw-bookings-not_logged_in"))

    def test_show_if_owner(self):
        self.booking_login()
        booking = self.create_booking()
        self.edit_place(booking)
        self.assertTextPresent("id_save_btn")

    def test_404_if_not_owner(self):
        self.booking_login()
        booking = self.create_booking()
        other_account = factories.create_booking_account()
        Booking.objects.all().update(account=other_account)
        self.edit_place(booking, expect_code=404)
        self.assertTextPresent("Page Not Found")

    def test_incomplete(self):
        self.booking_login()
        booking = self.create_booking()
        self.edit_place(booking)
        self.fill_by_name({"first_name": ""})
        self.submit_expecting_html5_validation_errors()
        self.assertTextPresent("This field is required")

    def test_complete(self):
        self.booking_login()
        booking = self.create_booking()
        self.edit_place(booking)
        data = self.get_place_details()
        data["first_name"] = "A New Name"
        self.fill_by_name(data)
        self.submit()
        self.assertUrlsEqual(reverse("cciw-bookings-basket_list_bookings"))

        # Did we alter it?
        booking.refresh_from_db()
        assert booking.first_name == "A New Name"

    def test_edit_booked(self):
        """
        Test we can't edit a booking when it is already booked.
        (or anything but BookingState.INFO_COMPLETE)
        """
        account = self.booking_login()
        self.create_booking()
        b = account.bookings.get()

        for state in [BookingState.BOOKED]:
            b.state = state
            b.save()

            # Check there is no save button
            self.edit_place(b)
            assert not self.is_element_present("#id_save_btn")
            # Check for message
            self.assertTextPresent("can only be changed by an admin.")

            # Attempt a post.

            # First, load a page with a working submit button:
            b.state = BookingState.INFO_COMPLETE
            b.save()
            self.edit_place(b)

            # Now change behind the scenes:
            b.state = state
            b.save()

            # Now submit
            data = self.get_place_details()
            data["first_name"] = "A New Name"
            self.fill_by_name(data)
            self.submit()
            # Check we didn't alter it
            assert account.bookings.get().first_name != "A New Name"


class TestEditPlaceWT(EditPlaceBase, WebTestBase):
    pass


class TestEditPlaceSL(EditPlaceBase, SeleniumBase):
    pass


def fix_autocomplete_fields(field_names: list[str]):
    """
    Returns a mixin class that adds some hacks so that auto-complete fields
    in the admin work with WebTest.
    """

    class FixAutocompleteFieldMixin:
        def fill_by_name(self, fields):
            new_fields = {}
            to_fix = []
            for field_name, value in fields.items():
                if field_name in field_names:
                    if self.is_full_browser_test:
                        # Fix later
                        to_fix.append((field_name, value))
                    else:
                        # Hack needed to cope with autocomplete widget and WebTest:
                        form, field, item = self._find_form_and_field_by_css_selector(
                            self.last_response,
                            f"[name={field_name}]",
                        )
                        # Modify the select widget so that it has the value we need
                        form.fields[field_name][0].options.append((str(value), False, ""))
                        new_fields[field_name] = value
                else:
                    new_fields[field_name] = value

            super().fill_by_name(new_fields)

            if self.is_full_browser_test:
                for field_name, value in to_fix:
                    # Hack to cope with autocomplete widget and Selenium
                    self.execute_script(
                        f"""django.jQuery('[name={field_name}]').append('<option value="{value}" selected="selected"></option>');"""
                    )

    return FixAutocompleteFieldMixin


class EditPlaceAdminBase(BookingBaseMixin, fix_autocomplete_fields(["account"]), CreateBookingWebMixin, FuncBaseMixin):
    def test_approve_and_unapprove(self):
        self.booking_login()
        booking = self.create_booking(serious_illness=True)

        self.officer_login(secretary := officers_factories.create_booking_secretary())

        self.get_url("admin:bookings_booking_change", booking.id)
        self.fill_by_name({"approvals-0-status": ApprovalStatus.APPROVED})
        self.submit("[name=_save]")

        booking = Booking.objects.get()
        approval: BookingApproval = booking.approvals.get()
        assert approval.is_approved
        assert approval.checked_by == secretary
        assert approval.checked_at is not None

        self.assertTextPresent(
            f"An email has been sent to {booking.account.email} telling them the place has been approved"
        )
        mails = mail.outbox
        assert len(mails) == 1
        assert mails[0].to == [booking.account.email]

        # Unapprove:
        self.get_url("admin:bookings_booking_change", booking.id)
        self.fill_by_name({"approvals-0-status": ApprovalStatus.DENIED})
        self.submit("[name=_save]")

        booking = Booking.objects.get()
        approval: BookingApproval = booking.approvals.get()
        assert not approval.is_approved

    def test_create(self):
        self.open_bookings()
        self.officer_login(officers_factories.create_booking_secretary())
        account = BookingAccount.objects.create(
            email=self.booker_email,
            name="Joe",
            address_post_code="XYZ",
        )
        self.get_url("admin:bookings_booking_add")
        fields = self.get_place_details()
        fields.update(
            {
                "account": account.id,
                "state": BookingState.BOOKED,
                "amount_due": "130.00",
                "manual_payment_amount": "100",
                "manual_payment_payment_type": str(ManualPaymentType.CHEQUE),
            }
        )
        self.fill_by_name(fields)
        self.submit("[name=_save]")
        self.assertTextPresent("Select booking")
        self.assertTextPresent("A confirmation email has been sent")
        booking = Booking.objects.get()
        assert not booking.created_online
        assert booking.account.manual_payments.count() == 1
        mp = booking.account.manual_payments.get()
        assert mp.payment_type == ManualPaymentType.CHEQUE
        assert mp.amount == Decimal("100")


class TestEditPlaceAdminWT(EditPlaceAdminBase, WebTestBase):
    pass


class TestEditPlaceAdminSL(EditPlaceAdminBase, SeleniumBase):
    pass


class EditAccountAdminBase(BookingBaseMixin, FuncBaseMixin):
    def test_create(self):
        self.officer_login(officers_factories.create_booking_secretary())
        self.get_url("admin:bookings_bookingaccount_add")
        self.fill_by_name(
            {
                "name": "Joe",
                "email": "joe@example.com",
                "address_post_code": "XYZ",
            }
        )
        self.submit("[name=_save]")
        self.assertTextPresent("was added successfully")
        account = BookingAccount.objects.get(email="joe@example.com")
        assert account.name == "Joe"

    def test_edit(self):
        account = factories.create_booking_account(name="Joe")
        account.manual_payments.create(
            amount=Decimal("10.00"),
            payment_type=ManualPaymentType.CHEQUE,
        )
        assert account.payments.count() == 1
        self.officer_login(officers_factories.create_booking_secretary())
        self.get_url("admin:bookings_bookingaccount_change", account.id)
        self.assertTextPresent("Payments")
        self.assertTextPresent("Payment: 10.00 from Joe via Cheque")
        self.fill_by_name({"name": "Mr New Name"})
        self.submit("[name=_save]")
        self.assertTextPresent("was changed successfully")
        account = refresh(account)
        assert account.name == "Mr New Name"


class TestEditAccountAdminWT(EditAccountAdminBase, WebTestBase):
    pass


class TestEditAccountAdminSL(EditAccountAdminBase, SeleniumBase):
    pass


class EditPaymentAdminBase(fix_autocomplete_fields(["account"]), BookingBaseMixin, FuncBaseMixin):
    def test_add_manual_payment(self):
        booking = factories.create_booking()
        self.officer_login(officers_factories.create_booking_secretary())
        account = booking.account
        self.get_url("admin:bookings_manualpayment_add")
        self.fill_by_name(
            {
                "account": account.id,
                "amount": "12.00",
            }
        )
        self.submit("[name=_save]")
        self.assertTextPresent("Manual payment of £12")
        self.assertTextPresent("was added successfully")
        account.refresh_from_db()
        assert account.manual_payments.count() == 1
        assert account.total_received == Decimal("12")


class TestEditPaymentAdminWT(EditPaymentAdminBase, WebTestBase):
    pass


class TestEditPaymentAdminSL(EditPaymentAdminBase, SeleniumBase):
    pass


class AccountTransferBase(fix_autocomplete_fields(["from_account", "to_account"]), FuncBaseMixin):
    def test_add_account_transfer(self):
        account_1 = BookingAccount.objects.create(email="account1@example.com", name="Joe")
        account_2 = BookingAccount.objects.create(email="account2@example.com", name="Jane")
        account_1.manual_payments.create(amount="100.00")
        account_1 = refresh(account_1)
        assert account_1.total_received == Decimal("100.00")

        assert account_1.payments.count() == 1

        self.officer_login(officers_factories.create_booking_secretary())

        self.get_url("admin:bookings_accounttransferpayment_add")
        self.fill_by_name(
            {
                "from_account": account_1.id,
                "to_account": account_2.id,
                "amount": "15",
            }
        )
        self.submit("[name=_save]")
        self.assertTextPresent("was added successfully")

        account_1 = refresh(account_1)
        account_2 = refresh(account_2)

        assert account_1.payments.count() == 2
        assert account_2.payments.count() == 1

        assert account_1.total_received == Decimal("85.00")
        assert account_2.total_received == Decimal("15.00")

        # Deleting causes more payments to restore the original value
        account_1.transfer_from_payments.get().delete()

        account_1 = refresh(account_1)
        account_2 = refresh(account_2)

        assert account_1.payments.count() == 3
        assert account_2.payments.count() == 2

        assert account_1.total_received == Decimal("100.00")
        assert account_2.total_received == Decimal("0.00")


class TestAccountTransferWT(AccountTransferBase, WebTestBase):
    pass


class TestAccountTransferSL(AccountTransferBase, SeleniumBase):
    pass


class ListBookingsBase(BookingBaseMixin, CreateBookingWebMixin, FuncBaseMixin):
    # This includes tests for most of the business logic

    urlname = "cciw-bookings-basket_list_bookings"

    def assert_book_button_enabled(self):
        assert self.is_element_present("#id_book_now_btn")
        assert not self.is_element_present("#id_book_now_btn[disabled]")

    def assert_book_button_disabled(self):
        assert self.is_element_present("#id_book_now_btn")
        assert self.is_element_present("#id_book_now_btn[disabled]")

    def enable_book_button(self):
        # Used for testing what happens if user enables button using browser
        # tools etc. i.e. checking that we have proper server side validation
        if self.is_full_browser_test:
            self.execute_script("""document.querySelector('#id_book_now_btn').removeAttribute('disabled')""")

    def test_redirect_if_not_logged_in(self):
        self.get_url(self.urlname)
        self.assertUrlsEqual(reverse("cciw-bookings-not_logged_in"))

    def test_show_bookings(self):
        self.booking_login()
        self.create_booking(name="Frédéric Bloggs")
        self.get_url(self.urlname)

        self.assertTextPresent("Camp Blue")
        self.assertTextPresent("Frédéric Bloggs")
        self.assertTextPresent("£100")
        self.assertTextPresent("No problems with this booking")
        self.assert_book_button_enabled()

    # TODO - maybe some of these should be moved to model based tests via get_booking_problems()

    def test_handle_custom_price(self):
        self.booking_login()
        self.create_booking(price_type=PriceType.CUSTOM)
        self.get_url(self.urlname)

        self.assertTextPresent("Camp Blue")
        self.assertTextPresent("Frédéric Bloggs")
        self.assertTextPresent("TBA")
        self.assertTextPresent("A custom discount needs to be arranged by the booking secretary")
        self.assert_book_button_disabled()
        self.assertTextPresent("This place cannot be booked for the reasons described above")

    def test_2nd_child_discount_allowed(self):
        self.booking_login()
        self.create_booking(price_type=PriceType.SECOND_CHILD)

        self.get_url(self.urlname)
        self.assertTextPresent(self.CANNOT_USE_2ND_CHILD)
        self.assert_book_button_disabled()

        # 2 places, both at 2nd child discount, is not allowed.
        self.create_booking(price_type=PriceType.SECOND_CHILD)

        self.get_url(self.urlname)
        self.assertTextPresent(self.CANNOT_USE_2ND_CHILD)
        self.assert_book_button_disabled()

    def test_2nd_child_discount_allowed_if_booked(self):
        """
        Test that we can have 2nd child discount if full price
        place is already booked.
        """
        account = self.booking_login()
        self.create_booking(first_name="Joe")
        account.bookings.update(state=BookingState.BOOKED)

        self.create_booking(first_name="Mary", price_type=PriceType.SECOND_CHILD)

        self.get_url(self.urlname)
        self.assert_book_button_enabled()

    def test_3rd_child_discount_allowed(self):
        self.booking_login()
        self.create_booking(price_type=PriceType.FULL)
        self.create_booking(price_type=PriceType.THIRD_CHILD)

        self.get_url(self.urlname)
        self.assertTextPresent("You cannot use a 3rd child discount")
        self.assert_book_button_disabled()

        # 3 places, with 2 at 3rd child discount, is not allowed.
        self.create_booking(price_type=PriceType.THIRD_CHILD)

        self.get_url(self.urlname)
        self.assertTextPresent("You cannot use a 3rd child discount")
        self.assert_book_button_disabled()

    def test_handle_serious_illness(self):
        self.booking_login()
        booking = self.create_booking(serious_illness=True)
        self.get_url(self.urlname)
        self.assertTextPresent("Must be approved by leader due to serious illness/condition")
        self.assert_book_button_disabled()
        assert booking in Booking.objects.need_approving()

    def test_minimum_age(self):
        # if born Aug 31st 2001, and thisyear == 2012, should be allowed on camp with
        # minimum_age == 11
        self.booking_login()
        self.create_booking(birth_date=date(year=self.camp.year - self.camp_minimum_age, month=8, day=31))
        self.get_url(self.urlname)
        self.assertTextAbsent(self.BELOW_MINIMUM_AGE)

        # if born 1st Sept 2001, and thisyear == 2012, should not be allowed on camp with
        # minimum_age == 11
        Booking.objects.all().delete()
        self.create_booking(birth_date=date(year=self.camp.year - self.camp_minimum_age, month=9, day=1))
        self.get_url(self.urlname)
        self.assertTextPresent(self.BELOW_MINIMUM_AGE)

    def test_maximum_age(self):
        # if born 1st Sept 2001, and thisyear == 2019, should be allowed on camp with
        # maximum_age == 17
        self.booking_login()
        self.create_booking(birth_date=date(year=self.camp.year - (self.camp_maximum_age + 1), month=9, day=1))
        self.get_url(self.urlname)
        self.assertTextAbsent(self.ABOVE_MAXIMUM_AGE)

        # if born Aug 31st 2001, and thisyear == 2019, should not be allowed on camp with
        # maximum_age == 17
        Booking.objects.all().delete()
        self.create_booking(birth_date=date(year=self.camp.year - (self.camp_maximum_age + 1), month=8, day=31))
        self.get_url(self.urlname)
        self.assertTextPresent(self.ABOVE_MAXIMUM_AGE)

    def test_no_places_left(self):
        for i in range(0, self.camp.max_campers):
            factories.create_booking(camp=self.camp, state=BookingState.BOOKED)

        self.booking_login()
        self.create_booking(sex="m")
        self.get_url(self.urlname)
        self.assertTextPresent(self.NO_PLACES_LEFT)
        self.assert_book_button_enabled()

        # Don't want a redundant message
        self.assertTextAbsent(self.NO_PLACES_LEFT_FOR_BOYS)

    def test_no_male_places_left(self):
        for i in range(0, self.camp.max_male_campers):
            factories.create_booking(camp=self.camp, sex="m", state=BookingState.BOOKED)

        self.booking_login()
        self.create_booking(sex="m")
        self.get_url(self.urlname)
        self.assertTextPresent(self.NO_PLACES_LEFT_FOR_BOYS)
        self.assert_book_button_enabled()

    def test_no_female_places_left(self):
        for i in range(0, self.camp.max_female_campers):
            factories.create_booking(camp=self.camp, sex="f", state=BookingState.BOOKED)

        self.booking_login()
        self.create_booking(sex="f")
        self.get_url(self.urlname)
        self.assertTextPresent(self.NO_PLACES_LEFT_FOR_GIRLS)
        self.assert_book_button_enabled()

    def test_not_enough_places_left(self):
        for i in range(0, self.camp.max_campers - 1):
            factories.create_booking(camp=self.camp, sex="m", state=BookingState.BOOKED)

        self.booking_login()
        self.create_booking(sex="f")
        self.create_booking(sex="f")
        self.get_url(self.urlname)
        self.assertTextPresent(self.NOT_ENOUGH_PLACES)
        self.assert_book_button_enabled()

    def test_not_enough_male_places_left(self):
        for i in range(0, self.camp.max_male_campers - 1):
            factories.create_booking(camp=self.camp, sex="m", state=BookingState.BOOKED)
        self.camp.bookings.update(state=BookingState.BOOKED)

        self.booking_login()
        self.create_booking(sex="m")
        self.create_booking(sex="m")
        self.get_url(self.urlname)
        self.assertTextPresent(self.NOT_ENOUGH_PLACES_FOR_BOYS)
        self.assert_book_button_enabled()

    def test_not_enough_female_places_left(self):
        for i in range(0, self.camp.max_female_campers - 1):
            factories.create_booking(camp=self.camp, sex="f", state=BookingState.BOOKED)
        self.camp.bookings.update(state=BookingState.BOOKED)

        self.booking_login()
        self.create_booking(sex="f")
        self.create_booking(sex="f")
        self.get_url(self.urlname)
        self.assertTextPresent(self.NOT_ENOUGH_PLACES_FOR_GIRLS)
        self.assert_book_button_enabled()

    def test_booking_after_closing_date(self):
        self.camp.last_booking_date = self.today - timedelta(days=1)
        self.camp.save()

        self.booking_login()
        self.create_booking()
        self.get_url(self.urlname)
        self.assertTextPresent(self.CAMP_CLOSED_FOR_BOOKINGS)
        self.assert_book_button_disabled()

    def test_handle_two_problem_bookings(self):
        # Test the error we get for more than one problem booking
        self.booking_login()
        self.create_booking(
            name="Frédéric Bloggs",
            price_type=PriceType.CUSTOM,
        )
        self.create_booking(name="Another Child", price_type=PriceType.CUSTOM)
        self.get_url(self.urlname)

        self.assertTextPresent("Camp Blue")
        self.assertTextPresent("Frédéric Bloggs")
        self.assertTextPresent("TBA")
        self.assertTextPresent("A custom discount needs to be arranged by the booking secretary")
        self.assert_book_button_disabled()
        self.assertTextPresent("These places cannot be booked for the reasons described above")

    def test_handle_mixed_problem_and_non_problem(self):
        # Test the message we get if one place is bookable and the other is not
        self.booking_login()
        self.create_booking()  # bookable
        self.create_booking(name="Another Child", price_type=PriceType.CUSTOM)  # not bookable
        self.get_url(self.urlname)
        self.assert_book_button_disabled()
        self.assertTextPresent("One or more of the places cannot be booked")

    def test_total(self):
        self.booking_login()
        self.create_booking()
        self.create_booking()
        self.get_url(self.urlname)
        self.assertTextPresent("£200")

    def test_manually_approved(self):
        # manually approved places should appear as OK to book
        self.booking_login()
        self.create_booking(name="Frédéric Bloggs")  # bookable
        booking2 = self.create_booking(name="Another Child", price_type=PriceType.CUSTOM)
        booking2.approve_booking_for_problem(type=ANT.CUSTOM_PRICE, user=officers_factories.create_booking_secretary())
        Booking.objects.filter(id=booking2.id).update(amount_due=Decimal("0.01"))
        self.get_url(self.urlname)

        self.assertTextPresent("Camp Blue")
        self.assertTextPresent("Frédéric Bloggs")
        self.assertTextPresent("£100")
        self.assertTextPresent("No problems with this booking")

        self.assertTextPresent("Another Child")
        self.assertTextPresent("£0.01")

        self.assert_book_button_enabled()
        # Total:
        self.assertTextPresent("£100.01")

    def test_add_another_btn(self):
        self.booking_login()
        self.create_booking()
        self.get_url(self.urlname)
        self.submit("[name=add_another]")
        self.assertUrlsEqual(reverse("cciw-bookings-add_place"))

    def test_book_ok(self):
        """
        Test that we can book a place
        """
        self.booking_login()
        booking = self.create_booking()
        self.get_url(self.urlname)
        self.submit("[name=book_now]")
        self.assertUrlsEqual(reverse("cciw-bookings-added_to_queue"))
        self.assertTextPresent("added to the queue")
        self.assertTextPresent("We will notify you by email whether or not these places have been allocated.")
        self.assertTextPresent(booking.name)

        booking.refresh_from_db()

        # Should be on queue, but not booked yet.
        assert booking.is_in_queue
        assert booking.state != BookingState.BOOKED

    def test_book_one_unbookable(self):
        """
        Test that if one places is unbookable, no place can be booked
        """
        account = self.booking_login()
        self.create_booking()
        self.create_booking(serious_illness=True)
        self.get_url(self.urlname)
        self.assert_book_button_disabled()
        self.enable_book_button()
        self.submit("[name=book_now]")
        for b in account.bookings.all():
            assert b.state == BookingState.INFO_COMPLETE
            assert not b.is_in_queue
        self.assertTextPresent("These places cannot be booked")

    def test_same_name_same_camp(self):
        self.booking_login()
        self.create_booking()
        self.create_booking()  # Identical

        self.get_url(self.urlname)
        self.assertTextPresent("You have entered another set of place details for a camper called")
        # This is only a warning:
        self.assert_book_button_enabled()

    def test_warn_about_multiple_full_price(self):
        self.booking_login()
        self.create_booking(name="Frédéric Bloggs")
        self.create_booking(name="Mary Bloggs")

        self.get_url(self.urlname)
        self.assertTextPresent(self.MULTIPLE_FULL_PRICE_WARNING)
        self.assertTextPresent("If Mary Bloggs and Frédéric Bloggs")
        # This is only a warning:
        self.assert_book_button_enabled()

        # Check for more than 2
        self.create_booking(name="Peter Bloggs")
        self.get_url(self.urlname)
        self.assertTextPresent("If Mary Bloggs, Peter Bloggs and Frédéric Bloggs")

    def test_warn_about_multiple_2nd_child(self):
        self.booking_login()
        self.create_booking(name="Frédéric Bloggs")
        self.create_booking(name="Mary Bloggs", price_type=PriceType.SECOND_CHILD)
        self.create_booking(name="Peter Bloggs", price_type=PriceType.SECOND_CHILD)

        self.get_url(self.urlname)
        self.assertTextPresent(self.MULTIPLE_2ND_CHILD_WARNING)
        self.assertTextPresent("If Peter Bloggs and Mary Bloggs")
        self.assertTextPresent("one is eligible")
        # This is only a warning:
        self.assert_book_button_enabled()

        self.create_booking(name="Zac Bloggs", price_type=PriceType.SECOND_CHILD)
        self.get_url(self.urlname)
        self.assertTextPresent("2 are eligible")

    def test_dont_warn_about_multiple_full_price_for_same_child(self):
        self.booking_login()
        self.create_booking()
        self.create_booking(camp=self.camp_2)

        self.get_url(self.urlname)
        self.assertTextAbsent(self.MULTIPLE_FULL_PRICE_WARNING)
        self.assert_book_button_enabled()

    def test_error_for_2nd_child_discount_for_same_camper(self):
        self.booking_login()
        self.create_booking()
        self.create_booking(camp=self.camp_2, price_type=PriceType.SECOND_CHILD)

        self.get_url(self.urlname)
        self.assertTextPresent(self.CANNOT_USE_2ND_CHILD)
        self.assert_book_button_disabled()

    def test_error_for_multiple_2nd_child_discount(self):
        self.booking_login()
        # Frederik x2
        self.create_booking(name="Peter Bloggs")
        self.create_booking(name="Peter Bloggs", camp=self.camp_2)

        # Mary x2
        self.create_booking(name="Mary Bloggs", price_type=PriceType.SECOND_CHILD)
        self.create_booking(name="Mary Bloggs", camp=self.camp_2, price_type=PriceType.SECOND_CHILD)

        self.get_url(self.urlname)
        self.assertTextPresent(self.CANNOT_USE_MULTIPLE_DISCOUNT_FOR_ONE_CAMPER)
        self.assert_book_button_disabled()

    def test_book_now_safeguard(self):
        self.booking_login()
        # It might be possible to alter the list of items in the basket in one
        # tab, and then press 'Book now' from an out-of-date representation of
        # the basket. We need a safeguard against this.

        # Must include at least id,price,camp choice for each booking
        booking = self.create_booking()
        self.get_url(self.urlname)

        # Now modify
        booking.refresh_from_db()
        booking.amount_due = Decimal("35.01")
        booking.save()

        self.submit("[name=book_now]")
        # Should not be modified
        booking.refresh_from_db()
        assert booking.state == BookingState.INFO_COMPLETE
        self.assertTextPresent("Places were not booked due to modifications made")


class TestListBookingsWT(ListBookingsBase, WebTestBase):
    pass


class TestListBookingsSL(ListBookingsBase, SeleniumBase):
    def test_move_to_shelf(self):
        self.booking_login()
        booking = self.create_booking()
        assert not booking.shelved
        self.get_url(self.urlname)

        self.click(f'tr[data-booking-id="{booking.id}"] [name=shelve]')
        self.wait_for_ajax()

        # Should be changed
        booking.refresh_from_db()
        assert booking.shelved

        # Different button should appear
        assert not self.is_element_present(f'tr[data-booking-id="{booking.id}"] [name=shelve]')
        assert self.is_element_present(f'tr[data-booking-id="{booking.id}"] [name=unshelve]')

    def test_move_to_basket(self):
        self.booking_login()
        booking = self.create_booking()
        booking.shelved = True
        booking.save()

        self.get_url(self.urlname)
        self.click(f'tr[data-booking-id="{booking.id}"] [name=unshelve]')
        self.wait_for_ajax()

        # Should be changed
        booking.refresh_from_db()
        assert not booking.shelved

    def test_delete_place(self):
        account = self.booking_login()
        booking = self.create_booking()
        self.get_url(self.urlname)

        self.click(f'tr[data-booking-id="{booking.id}"] [name=delete]', expect_alert=True)
        self.accept_alert()
        self.wait_for_ajax()

        # Should be gone
        if self.is_full_browser_test:
            self.wait_until(lambda d: account.bookings.count() == 0)
        else:
            assert account.bookings.count() == 0

    def test_edit_place_btn(self):
        self.booking_login()
        booking = self.create_booking()
        self.get_url(self.urlname)

        self.submit(f'tr[data-booking-id="{booking.id}"] [name=edit]')
        self.assertUrlsEqual(reverse("cciw-bookings-edit_place", kwargs={"booking_id": booking.id}))


class PayBase(BookingBaseMixin, CreateBookingWebMixin, FuncBaseMixin):
    def test_balance_empty(self):
        self.booking_login()
        self.open_bookings()
        self.get_url("cciw-bookings-pay")
        self.assertTextPresent("£0.00")

    def test_balance_after_booking(self):
        self.add_prices()
        self.booking_login()
        booking1 = self.create_booking()
        booking2 = self.create_booking()
        allocate_bookings_now([booking1, booking2])

        self.get_url("cciw-bookings-pay")

        # Move forward to after the time when full amount is required:
        Camp.objects.update(start_date=date.today() + timedelta(10))

        self.get_url("cciw-bookings-pay")

        # 2 full price
        expected_price = 2 * self.price_full
        self.assertTextPresent(f"£{expected_price}")


class TestPayWT(PayBase, WebTestBase):
    pass


class TestPaySL(PayBase, SeleniumBase):
    pass


class TestPayReturnPoints(BookingBaseMixin, BookingLogInMixin, WebTestBase):
    def test_pay_done(self):
        self.booking_login()
        self.get_url("cciw-bookings-pay_done")
        self.assertTextPresent("Payment complete!")
        # Paypal posts to these, check we support that
        resp = self.client.post(reverse("cciw-bookings-pay_done"), {})
        assert resp.status_code == 200

    def test_pay_cancelled(self):
        self.booking_login()
        self.get_url("cciw-bookings-pay_cancelled")
        self.assertTextPresent("Payment cancelled")
        # Paypal posts to these, check we support that
        resp = self.client.post(reverse("cciw-bookings-pay_cancelled"), {})
        assert resp.status_code == 200


@pytest.mark.django_db
def test_account_receive_payment():
    booking = factories.create_booking()
    (_, leader_1_user), (_, leader_2_user) = camps_factories.create_and_add_leaders(booking.camp, count=2)
    account = booking.account
    allocate_bookings_now([booking])
    booking.refresh_from_db()

    mail.outbox = []
    ManualPayment.objects.create(
        account=account,
        amount=booking.amount_due,
    )

    account.refresh_from_db()

    # Check we updated the account
    assert account.total_received == booking.amount_due
    assert account.total_received > 0  # sanity check

    # Check we updated the bookings
    booking.refresh_from_db()

    # Check for emails sent
    # 1 to account

    # TODO #52
    # mails = mail.outbox
    # account_mails = [m for m in mails if m.to == [account.email]]
    # assert len(account_mails) == 1


@pytest.mark.django_db
def test_email_for_bad_payment_1():
    ipn_1 = IpnMock()
    ipn_1.id = 123
    ipn_1.mc_gross = Decimal("1.00")
    ipn_1.custom = "x"  # wrong format

    mail.outbox = []
    assert len(mail.outbox) == 0
    paypal_payment_received(ipn_1)

    assert len(mail.outbox) == 1
    assert "/admin/ipn/paypal" in mail.outbox[0].body
    assert "No associated account" in mail.outbox[0].body


@pytest.mark.django_db
def test_email_for_bad_payment_2():
    account = BookingAccount(id=1234567)  # bad ID, not in DB
    ipn_1 = IpnMock()
    ipn_1.id = 123
    ipn_1.mc_gross = Decimal("1.00")
    ipn_1.custom = build_paypal_custom_field(account)

    mail.outbox = []
    assert len(mail.outbox) == 0
    paypal_payment_received(ipn_1)

    assert len(mail.outbox) == 1
    assert "/admin/ipn/paypal" in mail.outbox[0].body
    assert "No associated account" in mail.outbox[0].body


@pytest.mark.django_db
def test_email_for_bad_payment_3():
    ipn_1 = IpnMock()
    ipn_1.id = 123
    ipn_1.mc_gross = Decimal("1.00")

    mail.outbox = []
    assert len(mail.outbox) == 0
    unrecognised_payment(ipn_1)

    assert len(mail.outbox) == 1
    assert "/admin/ipn/paypal" in mail.outbox[0].body
    assert "Invalid IPN" in mail.outbox[0].body


@pytest.mark.django_db
def test_receive_payment_signal_handler():
    # Use the actual signal handler, check the good path.
    account = factories.create_booking_account()
    assert account.total_received == Decimal(0)

    ipn_1 = factories.create_ipn(account=account)

    # Test for Payment objects
    assert Payment.objects.count() == 1
    assert Payment.objects.all()[0].amount == ipn_1.mc_gross

    # Test account updated
    account.refresh_from_db()
    assert account.total_received == ipn_1.mc_gross

    # Test refund is wired up
    ipn_2 = factories.create_ipn(
        account=account,
        parent_txn_id="1",
        txn_id="2",
        mc_gross=-1 * ipn_1.mc_gross,
        payment_status="Refunded",
    )

    assert Payment.objects.count() == 2
    assert Payment.objects.order_by("-created_at")[0].amount == ipn_2.mc_gross

    account.refresh_from_db()
    assert account.total_received == Decimal(0)


@pytest.mark.django_db
def test_email_for_good_payment(mailoutbox):
    account = factories.create_booking_account()
    factories.create_ipn(account=account, mc_gross=Decimal(100))

    assert len(mailoutbox) == 1
    email = mailoutbox[0]

    assert email.subject == "[CCIW] Payment received"
    assert email.to == [account.email]
    assert "We have received your payment of £100" in email.body


@pytest.mark.django_db
def test_BookingAccount_concurrent_save():
    acc1 = BookingAccount.objects.create(email="foo@foo.com")
    acc2 = BookingAccount.objects.get(email="foo@foo.com")

    acc1.receive_payment(Decimal("100.00"))

    assert BookingAccount.objects.get(email="foo@foo.com").total_received == Decimal("100.00")

    acc2.save()  # this will have total_received = 0.00

    assert BookingAccount.objects.get(email="foo@foo.com").total_received == Decimal("100.00")


@pytest.mark.django_db
def test_pending_payment_handling():
    # This test is story-style - checks the whole process
    # of handling pending payments.

    # Create a place

    booking = factories.create_booking()
    account = booking.account

    # Book it
    allocate_bookings_now([booking])
    # Sanity check initial condition:
    mail.outbox = []
    booking.refresh_from_db()

    # Send payment that doesn't complete immediately
    ipn_1 = factories.create_ipn(
        account=account,
        txn_id="8DX10782PJ789360R",
        mc_gross=Decimal("20.00"),
        payment_status="Pending",
        pending_reason="echeck",
    )

    # Money should not be counted as received
    account = refresh(account)
    assert account.total_received == Decimal("0.00")

    # Custom email sent:
    assert len(mail.outbox) == 1
    m = mail.outbox[0]
    assert "We have received a payment of £20.00 that is pending" in m.body
    assert "echeck" in m.body

    # Check that we can tell the account has pending payments
    # and how much.
    three_days_later = timezone.now() + timedelta(days=3)
    assert account.get_pending_payment_total(now=three_days_later) == Decimal("20.00")

    # But pending payments are considered abandoned after 3 months.
    three_months_later = three_days_later + timedelta(days=30 * 3)
    assert account.get_pending_payment_total(now=three_months_later) == Decimal("0.00")

    # Once confirmed payment comes in, we consider that there are no pending payments.

    # A different payment doesn't affect whether pending ones are completed:
    factories.create_ipn(
        account=account,
        txn_id="ABCDEF123",  # DIFFERENT txn_id
        mc_gross=Decimal("10.00"),
        payment_status="Completed",
    )
    account = refresh(account)
    assert account.total_received == Decimal("10.00")
    assert account.get_pending_payment_total(now=three_days_later) == Decimal("20.00")

    # But the same TXN id is recognised as cancelling the pending payment
    factories.create_ipn(
        account=account,
        txn_id=ipn_1.txn_id,  # SAME txn_id
        mc_gross=ipn_1.mc_gross,
        payment_status="Completed",
    )
    account = refresh(account)
    assert account.total_received == Decimal("30.00")
    assert account.get_pending_payment_total(now=three_days_later) == Decimal("0.00")


class TestAjaxViews(BookingBaseMixin, CreateBookingWebMixin, WebTestBase):
    # Basic tests to ensure that the views that serve AJAX return something
    # sensible.

    # NB use a mixture of WebTest and Django client tests

    def _booking_problems_json(self, place_details):
        data = {}
        for k, v in place_details.items():
            data[k] = v.id if isinstance(v, models.Model) else v

        resp = self.client.post(reverse("cciw-officers-booking_problems_json"), data)
        return json.loads(resp.content.decode("utf-8"))

    def _initial_place_details(self):
        data = self.get_place_details()
        data["created_at_0"] = "1970-01-01"  # Simulate form, which doesn't supply created
        data["created_at_1"] = "00:00:00"
        return data

    def test_booking_problems(self):
        self.open_bookings()
        acc1 = BookingAccount.objects.create(email="foo@foo.com", address_post_code="ABC", name="Mr Foo")
        officer = officers_factories.create_booking_secretary()
        self.client.force_login(officer)
        resp = self.client.post(reverse("cciw-officers-booking_problems_json"), {"account": str(acc1.id)})

        assert resp.status_code == 200
        j = json.loads(resp.content.decode("utf-8"))
        assert not j["valid"]

        data = self._initial_place_details()
        data["account"] = str(acc1.id)
        data["state"] = BookingState.INFO_COMPLETE
        data["amount_due"] = "100.00"
        data["price_type"] = PriceType.CUSTOM
        j = self._booking_problems_json(data)
        assert j["valid"]
        assert "A custom discount needs to be arranged by the booking secretary" in j["problems"]

    def test_booking_problems_price_check(self):
        # Test that the price is checked.
        # This is a check that is only run for booking secretary
        self.open_bookings()
        acc1 = BookingAccount.objects.create(email="foo@foo.com", address_post_code="ABC", name="Mr Foo")
        officer = officers_factories.create_booking_secretary()
        self.client.force_login(officer)

        data = self._initial_place_details()
        data["account"] = str(acc1.id)
        data["state"] = BookingState.BOOKED
        data["amount_due"] = "0.00"
        data["price_type"] = PriceType.FULL
        j = self._booking_problems_json(data)
        assert any(
            p.startswith(f"The 'amount due' is not the expected value of £{self.price_full}") for p in j["problems"]
        )

    def test_booking_problems_full_refund(self):
        # Test that the price is checked.
        # This is a check that is only run for booking secretary
        self.open_bookings()
        acc1 = BookingAccount.objects.create(email="foo@foo.com", address_post_code="ABC", name="Mr Foo")
        officer = officers_factories.create_booking_secretary()
        self.client.force_login(officer)

        data = self._initial_place_details()
        data["account"] = str(acc1.id)

        # Check 'full refund' cancellation.
        data["state"] = BookingState.CANCELLED_FULL_REFUND
        data["amount_due"] = "20.00"
        data["price_type"] = PriceType.FULL
        j = self._booking_problems_json(data)
        assert any(p.startswith("The 'amount due' is not the expected value of £0.00") for p in j["problems"])


class AccountOverviewBase(BookingBaseMixin, CreateBookingWebMixin, FuncBaseMixin):
    urlname = "cciw-bookings-account_overview"

    def test_show(self):
        # Book a place and pay
        account = self.booking_login()
        booking1 = self.create_booking(name="Frédéric Bloggs")
        add_basket_to_queue([booking1], by_user=account)

        # Book another
        booking2 = self.create_booking(name="Another Child")
        add_basket_to_queue([booking2], by_user=account)

        # 3rd place, not booked at all
        self.create_booking(name="3rd Child")

        # 4th place, cancelled
        booking4 = self.create_booking(name="4th Child")
        booking4.state = BookingState.CANCELLED_FULL_REFUND
        booking4.auto_set_amount_due()
        booking4.save()

        self.get_url(self.urlname)

        # Another one, so that messages are cleared
        self.get_url(self.urlname)

        # Confirmed place
        self.assertTextPresent("Frédéric")

        # Booked place
        self.assertTextPresent("Another Child")

        # Basket/Shelf
        self.assertTextPresent("Basket and Saved for later")

        # TODO #52 - queue or not in queue

        # Cancellation
        self.assertTextPresent("Cancelled places")


class TestAccountOverviewWT(AccountOverviewBase, WebTestBase):
    pass


class TestAccountOverviewSL(AccountOverviewBase, SeleniumBase):
    def test_manage_place_withdraw(self):
        account = self.booking_login()
        booking = self.create_booking(shortcut=True)
        booking.add_to_queue(by_user=account)

        booking.refresh_from_db()
        assert booking.is_in_queue

        self.get_url(self.urlname)
        manage_button_selector = f"#id_manage_booking_button_{booking.id}"
        self.click(manage_button_selector)
        self.wait_for_ajax()
        self.click("button[name=withdraw]")
        self.wait_for_ajax()

        booking.refresh_from_db()
        assert not booking.is_in_queue

        assert not self.is_element_present(manage_button_selector)

        # We should have an action log
        logs = list(booking.queue_entry.action_logs.all())
        last_log = logs[-1]
        assert last_log.account_user == account
        assert last_log.action_type == QueueEntryActionLogType.FIELDS_CHANGED
        assert last_log.details["fields_changed"] == [{"name": "is_active", "old_value": True, "new_value": False}]


class LogOutBase(BookingBaseMixin, BookingLogInMixin, FuncBaseMixin):
    def test_logout(self):
        self.booking_login()
        self.get_url("cciw-bookings-account_overview")
        self.submit("[name=logout]")
        self.assertUrlsEqual(reverse("cciw-bookings-index"))

        # Try accessing a page which is restricted
        self.get_url("cciw-bookings-account_overview")
        self.assertUrlsEqual(reverse("cciw-bookings-not_logged_in"))


class TestLogOutWT(LogOutBase, WebTestBase):
    pass


class TestLogOutSL(LogOutBase, SeleniumBase):
    pass


@pytest.mark.django_db
def test_ManualPayment_create():
    account = BookingAccount.objects.create(email="foo@foo.com")
    assert Payment.objects.count() == 0
    ManualPayment.objects.create(account=account, amount=Decimal("100.00"))
    assert Payment.objects.count() == 1
    assert Payment.objects.all()[0].amount == Decimal("100.00")

    account = BookingAccount.objects.get(id=account.id)
    assert account.total_received == Decimal("100.00")


@pytest.mark.django_db
def test_ManualPayment_delete():
    # Setup
    account = BookingAccount.objects.create(email="foo@foo.com")
    mp = ManualPayment.objects.create(account=account, amount=Decimal("100.00"))
    assert Payment.objects.count() == 1

    # Test
    mp.delete()
    assert Payment.objects.count() == 2
    account = BookingAccount.objects.get(id=account.id)
    assert account.total_received == Decimal("0.00")


@pytest.mark.django_db
def test_ManualPayment_edit():
    # Setup
    account = BookingAccount.objects.create(email="foo@foo.com")
    mp = ManualPayment.objects.create(account=account, amount=Decimal("100.00"))

    mp.amount = Decimal("101.00")
    with pytest.raises(Exception):
        mp.save()


@pytest.mark.django_db
def test_RefundPayment_create():
    account = BookingAccount.objects.create(email="foo@foo.com")
    assert Payment.objects.count() == 0
    RefundPayment.objects.create(account=account, amount=Decimal("100.00"))
    assert Payment.objects.count() == 1
    assert Payment.objects.all()[0].amount == Decimal("-100.00")

    account = BookingAccount.objects.get(id=account.id)
    assert account.total_received == Decimal("-100.00")


@pytest.mark.django_db
def test_RefundPayment_delete():
    # Setup
    account = BookingAccount.objects.create(email="foo@foo.com")
    rp = RefundPayment.objects.create(account=account, amount=Decimal("100.00"))
    assert Payment.objects.count() == 1

    # Test
    rp.delete()
    assert Payment.objects.count() == 2
    account = BookingAccount.objects.get(id=account.id)
    assert account.total_received == Decimal("0.00")


@pytest.mark.django_db
def test_RefundPayment_edit():
    # Setup
    account = BookingAccount.objects.create(email="foo@foo.com")
    rp = RefundPayment.objects.create(account=account, amount=Decimal("100.00"))

    rp.amount = Decimal("101.00")
    with pytest.raises(Exception):
        rp.save()


@pytest.mark.django_db
def test_cancel_amount_due():
    booking = factories.create_booking()
    booking.state = BookingState.CANCELLED_FULL_REFUND
    assert booking.expected_amount_due() == Decimal(0)


@pytest.mark.django_db
def test_cancel_account_amount_due():
    booking = factories.create_booking()
    account = booking.account
    booking.state = BookingState.CANCELLED_HALF_REFUND
    assert booking.expected_amount_due() > Decimal(0)
    booking.auto_set_amount_due()
    booking.save()

    account.refresh_from_db()
    assert account.get_balance_full() == booking.amount_due


@pytest.mark.django_db
def test_cancel_full_refund_amount_due():
    booking = factories.create_booking()
    booking.state = BookingState.CANCELLED_FULL_REFUND
    assert booking.expected_amount_due() == Decimal("0.00")


@pytest.mark.django_db
def test_cancel_half_refund_amount_due():
    booking = factories.create_booking()
    booking.state = BookingState.CANCELLED_HALF_REFUND
    assert (
        booking.expected_amount_due() == Price.objects.get(year=booking.camp.year, price_type=PriceType.FULL).price / 2
    )


@pytest.mark.django_db
def test_cancel_booking_fee_kept_amount_due():
    booking = factories.create_booking()
    booking.state = BookingState.CANCELLED_BOOKING_FEE_KEPT

    assert (
        booking.expected_amount_due()
        == Price.objects.get(year=booking.camp.year, price_type=PriceType.BOOKING_FEE).price
    )


@pytest.mark.django_db
def test_cancel_full_refund_account_amount_due():
    booking = factories.create_booking()
    account = booking.account
    booking.state = BookingState.CANCELLED_FULL_REFUND
    booking.auto_set_amount_due()
    booking.save()

    account.refresh_from_db()
    assert account.get_balance_full() == booking.amount_due


@pytest.mark.django_db
def test_export_places_summary():
    booking = factories.create_booking()
    booking.state = BookingState.BOOKED
    booking.save()

    workbook = camp_bookings_to_spreadsheet(booking.camp).to_bytes()
    wkbk: openpyxl.Workbook = openpyxl.load_workbook(io.BytesIO(workbook))

    wksh_all = wkbk.worksheets[0]

    assert wksh_all.cell(1, 1).value == "First name"
    assert wksh_all.cell(2, 1).value == booking.first_name


@pytest.mark.django_db
@pytest.mark.parametrize(
    # Choose some different dates to set camp, to exercise logic for calculating
    # birthdays. Birthday is set day after camp start date.
    "dates_and_age",
    [
        # camp start date, birth date, age on camp, birthday on camp
        (
            # Day after camp starts
            date(2012, 8, 1),
            date(2000, 8, 2),
            12,
            date(2012, 8, 2),
        ),
        (
            # Day camp starts
            date(2012, 8, 1),
            date(2000, 8, 1),
            12,
            date(2012, 8, 1),
        ),
        (
            # Leap years!
            date(2017, 2, 28),  # not leap year
            date(2004, 2, 29),  # leap birthday
            13,
            date(2017, 2, 28),  # celebrate the day before
        ),
        (
            # Edge case for (unrealistic) scenario where camp spans a year end.
            date(2025, 12, 31),
            date(2013, 1, 1),
            13,
            date(2026, 1, 1),
        ),
    ],
)
def test_export_places_birthdays(dates_and_age: tuple[date, date, int, date]):
    camp = camps_factories.create_camp(start_date=dates_and_age[0])
    birth_date = dates_and_age[1]
    age = dates_and_age[2]
    birthday_on_camp = dates_and_age[3]
    booking = factories.create_booking(birth_date=birth_date, camp=camp, state=BookingState.BOOKED)

    workbook = camp_bookings_to_spreadsheet(camp).to_bytes()
    wkbk: openpyxl.Workbook = openpyxl.load_workbook(io.BytesIO(workbook))
    wksh_bdays = wkbk.worksheets[2]

    assert wksh_bdays.cell(1, 1).value == "First name"
    assert wksh_bdays.cell(2, 1).value == booking.first_name

    assert wksh_bdays.cell(1, 3).value == "Birthday"
    assert wksh_bdays.cell(2, 3).value == birthday_on_camp.strftime("%A %d %B")

    assert wksh_bdays.cell(1, 4).value == "Age"
    assert wksh_bdays.cell(2, 4).value == str(age)


@pytest.mark.django_db
def test_export_payment_data():
    account1 = BookingAccount.objects.create(name="Joe Bloggs", email="joe@foo.com")
    account2 = BookingAccount.objects.create(name="Mary Muddle", email="mary@foo.com")
    factories.create_ipn(account=account1, mc_gross=Decimal("10.00"))
    ManualPayment.objects.create(account=account1, amount=Decimal("11.50"))
    RefundPayment.objects.create(account=account1, amount=Decimal("0.25"))
    AccountTransferPayment.objects.create(from_account=account2, to_account=account1, amount=Decimal("100.00"))
    mp2 = ManualPayment.objects.create(account=account1, amount=Decimal("1.23"))
    mp2.delete()

    now = timezone.now()
    workbook = payments_to_spreadsheet(now - timedelta(days=3), now + timedelta(days=3)).to_bytes()

    wkbk: openpyxl.Workbook = openpyxl.load_workbook(io.BytesIO(workbook))
    wksh = wkbk.worksheets[0]
    data = [[c.value for c in r] for r in wksh.rows]
    assert data[0] == ["Account name", "Account email", "Amount", "Date", "Type"]

    # Excel dates are a pain, so we ignore them
    data2 = [[c for i, c in enumerate(r) if i != 3] for r in data[1:]]
    assert ["Joe Bloggs", "joe@foo.com", 10.0, "PayPal"] in data2
    assert ["Joe Bloggs", "joe@foo.com", 11.5, "Cheque"] in data2
    assert ["Joe Bloggs", "joe@foo.com", -0.25, "Refund Cheque"] in data2
    assert ["Joe Bloggs", "joe@foo.com", 100.00, "Account transfer"] in data2

    assert ["Joe Bloggs", "joe@foo.com", 1.23, "ManualPayment (deleted)"] not in data2
    assert ["Joe Bloggs", "joe@foo.com", -1.23, "ManualPayment (deleted)"] not in data2


@pytest.mark.django_db
def test_booking_saved_approvals_unapproved_and_need_approving():
    booking = factories.create_booking()
    assert len(Booking.objects.need_approving()) == 0

    booking.serious_illness = True
    booking.birth_date = date(1980, 1, 1)
    booking.price_type = PriceType.CUSTOM
    booking.save()
    booking.update_approvals()

    assert len(Booking.objects.need_approving()) == 1
    booking: Booking = Booking.objects.get()
    camper_age = booking.age_on_camp()
    camp = booking.camp
    actual_approvals = [(app.description, app.type) for app in booking.saved_approvals_unapproved]
    expected_approvals = [
        ("A custom discount needs to be arranged by the booking secretary", ANT.CUSTOM_PRICE),
        ("Must be approved by leader due to serious illness/condition", ANT.SERIOUS_ILLNESS),
        (
            f"Camper will be {camper_age} which is above the maximum age ({camp.maximum_age}) on 31 August {camp.year}",
            ANT.TOO_OLD,
        ),
    ]
    assert actual_approvals == expected_approvals
    assert booking.saved_approvals_needed_summary == "Custom price, Serious illness, Too old"

    # Check that update_approvals adds and removes correctly.
    booking.serious_illness = False
    booking.birth_date = date(camp.year - 2, 1, 1)
    booking.price_type = PriceType.FULL
    booking.save()
    booking.update_approvals()

    booking = Booking.objects.get()
    types = [app.type for app in booking.saved_approvals_unapproved]
    assert types == [ANT.TOO_YOUNG]

    # Check that `need_approving` responds to approvals being done.
    assert len(Booking.objects.need_approving()) == 1

    booking.approve_booking_for_problem(type=ANT.TOO_YOUNG, user=officers_factories.get_any_officer())

    assert len(Booking.objects.need_approving()) == 0

    # Check that approve_booking_for_problem actually worked
    booking = Booking.objects.get()
    assert booking.saved_approvals_unapproved == []


@pytest.mark.django_db
def test_booking_add_to_queue():
    booking = factories.create_booking()
    assert not booking.is_in_queue
    booking.add_to_queue(by_user=booking.account)
    assert booking.is_in_queue
    booking.refresh_from_db()
    assert booking.queue_entry.is_active
    action_log = booking.queue_entry.action_logs.all()[0]
    assert action_log.account_user == booking.account
    assert action_log.staff_user is None
    assert action_log.action_type == QueueEntryActionLogType.CREATED

    # Multiple times is fine and does nothing.
    old_queue_entry = booking.queue_entry
    created_at = old_queue_entry.created_at

    booking.add_to_queue(by_user=booking.account)
    booking.refresh_from_db()

    assert booking.queue_entry == old_queue_entry
    assert booking.queue_entry.created_at == created_at


@pytest.mark.django_db
def test_payment_source_save_bad():
    manual = factories.create_manual_payment()
    refund = factories.create_refund_payment()
    with pytest.raises(AssertionError):
        PaymentSource.objects.create(manual_payment=manual, refund_payment=refund)


@pytest.mark.django_db
def test_payment_source_save_good():
    manual = factories.create_manual_payment()
    PaymentSource.objects.all().delete()
    p = PaymentSource.objects.create(manual_payment=manual)
    assert p.id is not None


@pytest.mark.django_db
def test_write_off_debt_payment():
    account = factories.create_booking_account()
    factories.create_booking(account=account, state=BookingState.BOOKED)
    account.refresh_from_db()

    balance = account.get_balance_full()
    assert balance > 0

    factories.create_write_off_debt_payment(account=account, amount=balance)
    account.refresh_from_db()

    assert account.get_balance_full() == 0


class SupportingInformationAdminBase(fix_autocomplete_fields("booking"), FuncBaseMixin):
    def test_separate_supporting_information_admin(self):
        booking = factories.create_booking()
        information_type = factories.create_supporting_information_type(name="test")
        self.officer_login(officers_factories.create_booking_secretary())
        self.get_url("admin:bookings_supportinginformation_add")
        self.fill_by_name(
            {
                "booking": booking.id,
                "information_type": information_type.id,
                "from_name": "Zog",
                "from_email": "zog@example.com",
                "from_telephone": "1234 567890",
                "notes": "These are some notes",
                "document": Upload("hello.txt", b"Hello"),
            }
        )
        self.submit("[name=_continue]")
        self.assertTextPresent("was added successfully")
        supporting_information = booking.supporting_information_records.get()
        assert supporting_information.information_type == information_type
        assert supporting_information.from_name == "Zog"
        assert supporting_information.from_email == "zog@example.com"
        assert supporting_information.from_telephone == "1234 567890"
        assert supporting_information.notes == "These are some notes"
        doc = supporting_information.document
        assert doc.filename.endswith("hello.txt")  # functest limitation means we don't get exact name
        assert bytes(doc.content) == b"Hello"
        assert doc.size == 5
        assert doc.mimetype == "text/plain"

        # Save again, without upload, shouldn't clear the doc.
        self.fill_by_name({"from_name": "Zog2"})
        self.submit("[name=_continue]")
        self.assertTextPresent("was changed successfully")
        supporting_information.refresh_from_db()
        assert supporting_information.from_name == "Zog2"
        assert supporting_information.document is not None
        assert supporting_information.document.filename.endswith("hello.txt")

        # Test clear
        self.fill_by_name({"document-clear": True})
        self.submit("[name=_continue]")
        self.assertTextPresent("was changed successfully")
        supporting_information.refresh_from_db()
        assert supporting_information.document is None
        self.assertTextPresent("was changed successfully")

        # Saving without upload (or file initially attached) should be allowed
        self.fill_by_name({"from_name": "Zog3"})
        self.submit("[name=_continue]")
        self.assertTextPresent("was changed successfully")
        supporting_information.refresh_from_db()
        assert supporting_information.from_name == "Zog3"

    def test_invalid_form_with_new_upload(self):
        supporting_information = factories.create_supporting_information(document_filename="old.txt")
        self.officer_login(officers_factories.create_booking_secretary())
        self.get_url("admin:bookings_supportinginformation_change", object_id=supporting_information.id)
        self.fill_by_name(
            {
                "from_name": "",  # Invalid, should block save
                "document": Upload("new_doc.txt", b"Hello"),
            }
        )
        self.submit("[name=_continue]")
        self.assertTextAbsent("was changed successfully")
        supporting_information.refresh_from_db()
        assert supporting_information.document.filename == "old.txt"

    def test_invalid_form_with_clear_upload(self):
        supporting_information = factories.create_supporting_information(document_filename="old.txt")
        self.officer_login(officers_factories.create_booking_secretary())
        self.get_url("admin:bookings_supportinginformation_change", object_id=supporting_information.id)
        self.fill_by_name(
            {
                "from_name": "",  # Invalid, should block save
                "document-clear": True,
            }
        )
        self.submit("[name=_continue]")
        self.assertTextAbsent("was changed successfully")
        supporting_information.refresh_from_db()
        assert supporting_information.document is not None
        assert supporting_information.document.filename == "old.txt"

    def test_supporting_information_inline(self):
        booking = factories.create_booking()
        information_type = factories.create_supporting_information_type(name="test")
        self.officer_login(officers_factories.create_booking_secretary())
        self.get_url("admin:bookings_booking_change", booking.id)
        if self.is_full_browser_test:
            self.click("#supporting_information_records-group details")
            self.click("#supporting_information_records-group .add-row a")
        else:
            self.add_admin_inline_form_to_page("supporting_information_records")
        self.fill_by_name(
            {
                f"supporting_information_records-0-{k}": v
                for k, v in {
                    "information_type": information_type.id,
                    "from_name": "Zog",
                    "document": Upload("hello.txt", b"Hello"),
                }.items()
            }
        )
        self.submit("[name=_continue]")
        supporting_information = booking.supporting_information_records.get()
        # For other fields tests are above, we care most about file upload, which is
        # trickiest
        doc = supporting_information.document
        assert bytes(doc.content) == b"Hello"

        # Test clear
        if self.is_full_browser_test:
            self.click("#supporting_information_records-group details")
        self.fill_by_name({"supporting_information_records-0-document-clear": True})
        self.submit("[name=_continue]")
        self.assertTextPresent("was changed successfully")
        supporting_information.refresh_from_db()
        assert supporting_information.document is None


class SupportingInformationAdminWT(SupportingInformationAdminBase, WebTestBase):
    pass


class SupportingInformationAdminSL(SupportingInformationAdminBase, SeleniumBase):
    pass


class DocumentDownloadView(WebTestBase):
    def test_deny_normal_user(self):
        info = factories.create_supporting_information(document_filename="anything.txt")
        self.officer_login(officers_factories.create_officer())
        response = self.get_literal_url(info.document.url, expect_errors=True)
        assert response.status_code == 404

    def test_allow_for_authorised_user(self):
        info = factories.create_supporting_information(
            document_filename="temp.txt",
            document_content=b"Hello",
            document_mimetype="text/plain",
        )
        self.officer_login(officers_factories.create_booking_secretary())
        response = self.get_literal_url(info.document.url)
        assert response.status_code == 200
        assert response.content == b"Hello"
        assert response.headers["Content-Type"] == "text/plain"
        assert response.headers["Content-Disposition"] == 'attachment; filename="temp.txt"'


@given(st.emails())
def test_decode_inverts_encode(email):
    v = EmailVerifyTokenGenerator()
    assert v.email_from_token(v.token_for_email(email)) == email


@given(st.emails())
def test_truncated_returns_invalid(email):
    v = EmailVerifyTokenGenerator()
    assert isinstance(v.email_from_token(v.token_for_email(email)[2:]), VerifyFailed)


@given(st.emails())
def test_expired_returns_expired(email):
    v = EmailVerifyTokenGenerator()
    assert v.email_from_token(v.token_for_email(email), max_age=-1) == VerifyExpired(email)


@given(email=st.text())
@example(email="abcdefgh")  # b64 encode results in trailing ==
def test_tolerate_truncated_trailing_equals(email):
    v = EmailVerifyTokenGenerator()

    # Either some silly people, or some dumb email programs, decide to strip
    # trailing = from URLs (despite this being a supposedly URL safe
    # character). Ensure that we tolerate this.

    def remove_equals(s):
        return s.rstrip("=")

        assert v.email_from_token(remove_equals(v.token_for_email(email))) == email


@pytest.mark.django_db
def test_booking_open():
    # Initially:
    year = date.today().year  # doesn't really matter
    assert not get_booking_open_data(year).is_open_for_booking
    assert not get_booking_open_data(year).is_open_for_entry

    factories.create_prices(year=year)
    assert not get_booking_open_data(year).is_open_for_booking
    assert not get_booking_open_data(year).is_open_for_entry

    config = factories.create_year_config(
        year=year, bookings_open_for_booking_on="future", bookings_open_for_entry_on="future"
    )
    assert not get_booking_open_data(year).is_open_for_booking
    assert not get_booking_open_data(year).is_open_for_entry
    config.delete()

    config = factories.create_year_config(
        year=year, bookings_open_for_booking_on="past", bookings_open_for_entry_on="past"
    )
    assert get_booking_open_data(year).is_open_for_booking
    assert get_booking_open_data(year).is_open_for_entry
    config.delete()

    config = factories.create_year_config(
        year=year, bookings_open_for_booking_on="future", bookings_open_for_entry_on="past"
    )
    assert not get_booking_open_data(year).is_open_for_booking
    assert get_booking_open_data(year).is_open_for_entry

    # Deleting prices closes booking
    Price.objects.all().delete()
    assert not get_booking_open_data(year).is_open_for_booking
    assert not get_booking_open_data(year).is_open_for_entry


def create_year_config_for_queue_tests(year: int = 2026) -> YearConfig:
    # Use sensible values that match what happens in reality,
    # and the defaults in `create_camp()` factory
    return factories.create_year_config(
        year=year,
        bookings_open_for_entry_on=date(year, 2, 1),
        bookings_open_for_booking_on=date(year, 3, 1),
        bookings_close_for_initial_period_on=date(year, 4, 1),
    )


@pytest.mark.django_db
def test_rank_queue_booking():
    year_config = create_year_config_for_queue_tests(year=2026)
    with freeze_time(year_config.bookings_open_for_entry_on + timedelta(days=1)):
        b1 = factories.create_booking(first_name="Amy")
        b2 = factories.create_booking(first_name="Bob")
        b3 = factories.create_booking(first_name="Carla")
        b4 = factories.create_booking(first_name="Dave")
        b5 = factories.create_booking(first_name="Ed")
    with freeze_time(year_config.bookings_open_for_booking_on + timedelta(days=1)):
        b1.add_to_queue(by_user=b1.account)
    with freeze_time(year_config.bookings_open_for_booking_on + timedelta(days=2)):
        b2.add_to_queue(by_user=b2.account)
        assert date.today() < year_config.bookings_close_for_initial_period_on
    with freeze_time(year_config.bookings_close_for_initial_period_on + timedelta(days=1)):
        assert date.today() > year_config.bookings_close_for_initial_period_on
        b3.add_to_queue(by_user=b3.account)

    with freeze_time(year_config.bookings_close_for_initial_period_on + timedelta(days=2)):
        b4.add_to_queue(by_user=b4.account)

    ranked_bookings = rank_queue_bookings(camp=b1.camp, year_config=year_config)

    assert b1 in ranked_bookings
    assert b2 in ranked_bookings
    assert b3 in ranked_bookings
    assert b4 in ranked_bookings
    assert b5 not in ranked_bookings  # never added to queue

    # Don't know if b1 or b2 will be first.
    b1_q = [b for b in ranked_bookings if b.id == b1.id][0]
    b2_q = [b for b in ranked_bookings if b.id == b2.id][0]

    b3_q = ranked_bookings[2]
    assert b3_q.id == b3.id

    b4_q = ranked_bookings[3]
    assert b4_q.id == b4.id

    # First two are first equal, due to being within the initial period
    assert b1_q.rank_info.queue_position_rank == 1
    assert b2_q.rank_info.queue_position_rank == 1

    # b3 and b4 are after the cut-off
    assert b3_q.rank_info.queue_position_rank == 2
    assert b4_q.rank_info.queue_position_rank == 3


@pytest.mark.django_db
def test_rank_queue_booking_same_camper_multiple_camps():
    year = 2026
    year_config = create_year_config_for_queue_tests(year=year)
    camp_1 = camps_factories.create_camp(year=year)
    camp_2 = camps_factories.create_camp(year=year)
    with freeze_time(year_config.bookings_open_for_entry_on + timedelta(days=1)):
        b1 = factories.create_booking(camp=camp_1, first_name="Amy")
        b2 = factories.create_booking(camp=camp_2, first_name=b1.first_name, last_name=b1.last_name, account=b1.account)
        b1.add_to_queue(by_user=b1.account)
        b2.add_to_queue(by_user=b2.account)
    with freeze_time(year_config.bookings_close_for_initial_period_on + timedelta(days=1)):
        b3 = factories.create_booking(camp=camp_2, first_name="Bob")
        assert date.today() > year_config.bookings_close_for_initial_period_on
        b3.add_to_queue(by_user=b3.account)

    # Before we book, we just know that b1 (and b2) have other places in queue:
    ranked_bookings_camp_1 = rank_queue_bookings(camp=camp_1, year_config=year_config)
    assert (b1_q := ranked_bookings_camp_1[0]) == b1
    assert b1_q.rank_info.has_other_place_in_queue
    assert not b1_q.rank_info.has_other_place_booked

    # Then we booked one of them:
    allocate_bookings_now([b1])

    # And rank the other:
    ranked_bookings_camp_2 = rank_queue_bookings(camp=camp_2, year_config=year_config)

    assert b2 in ranked_bookings_camp_2
    assert b3 in ranked_bookings_camp_2

    # b2 will be at bottom, despite b3 being after the initial period...
    assert (b3_q := ranked_bookings_camp_2[0]) == b3
    assert (b2_q := ranked_bookings_camp_2[1]) == b2

    # ...due to having another place booked
    assert b2_q.rank_info.has_other_place_booked
    assert not b3_q.rank_info.has_other_place_booked

    # We also track if camper has other place in queue, to display in UI
    assert b1_q.rank_info.has_other_place_in_queue
    assert b2_q.rank_info.has_other_place_in_queue
    assert not b3_q.rank_info.has_other_place_in_queue


@pytest.mark.django_db
def test_Booking_withdraw_from_queue_and_add_again():
    with freeze_time("2026-01-01"):
        booking = factories.create_booking()
        booking.add_to_queue(by_user=booking.account)
        booking.refresh_from_db()
        queue_entry_id = booking.queue_entry.id
    with freeze_time("2026-01-02"):
        booking.withdraw_from_queue(by_user=booking.account)
        booking.refresh_from_db()
        assert not booking.queue_entry.is_active
    with freeze_time("2026-01-03"):
        booking.add_to_queue(by_user=booking.account)
        booking.refresh_from_db()
        queue_entry_id2 = booking.queue_entry.id
        # For auditing, it's simpler if we keep old BookingQueueEntry:
        assert queue_entry_id == queue_entry_id2
        assert booking.queue_entry.is_active
        # But we need the `enqueued_at` to be updated
        # to the date they were re-added to the queue.
        assert booking.queue_entry.enqueued_at.date() == date(2026, 1, 3)

    assert booking.queue_entry.action_logs.count() == 3


@pytest.mark.django_db
def test_QueueEntry_get_current_field_data():
    booking = factories.create_booking()
    queue_entry = booking.add_to_queue(by_user=booking.account)
    keys = sorted(queue_entry.get_current_field_data().keys())
    # Avoid this breaking for each new field by testing a subset
    assert all(
        k in keys
        for k in [
            "booking_id",
            "created_at",
            "enqueued_at",
            "erased_at",
            "first_timer_allocated",
            "id",
            "is_active",
            "officer_child",
            "sibling_booking_account_id",
            "sibling_fuzzy_id",
            "sibling_surname",
        ]
    )


@pytest.mark.django_db
def test_get_booking_queue_problems_rejected_first_timers():
    year_config = create_year_config_for_queue_tests()
    camp = camps_factories.create_camp(
        year=year_config.year, max_campers=20, max_male_campers=10, max_female_campers=10
    )
    bookings: list[Booking] = [factories.create_booking(camp=camp, sex=Sex.MALE) for _ in range(0, 15)]
    for b in bookings:
        queue_entry = b.add_to_queue(by_user=b.account)
        queue_entry.first_timer_allocated = True
        queue_entry.save()
    ranked_queue_bookings = rank_queue_bookings(camp=camp, year_config=year_config)
    add_queue_cutoffs(ranked_queue_bookings=ranked_queue_bookings, places_left=camp.get_places_left())
    problems = get_booking_queue_problems(ranked_queue_bookings=ranked_queue_bookings, camp=camp)
    assert len(problems.rejected_first_timers) == 5


@pytest.mark.django_db
def test_get_booking_queue_problems_too_many_first_timers():
    year_config = create_year_config_for_queue_tests()
    camp = camps_factories.create_camp(
        year=year_config.year, max_campers=20, max_male_campers=10, max_female_campers=10
    )
    bookings: list[Booking] = [factories.create_booking(camp=camp, sex=Sex.MALE) for _ in range(0, 10)]
    for b in bookings:
        queue_entry = b.add_to_queue(by_user=b.account)
        queue_entry.first_timer_allocated = True
        queue_entry.save()
    problems = get_booking_queue_problems(ranked_queue_bookings=[], camp=camp)
    assert (
        problems.general_messages[0] == '10 bookings are marked as "chosen first timers", but only 2 are allowed (10%)'
    )


class BookingQueuePageBase(FuncBaseMixin):
    def _ensure_camp(self):
        if not hasattr(self, "year_config"):
            self.year_config = create_year_config_for_queue_tests()
        if not hasattr(self, "camp"):
            self.camp: Camp = camps_factories.create_camp(year=self.year_config.year)

    def _create_booking(self, first_name: str = Auto) -> Booking:
        self._ensure_camp()
        return factories.create_booking(camp=self.camp, first_name=first_name)

    def test_allocate_places(self):
        COUNT = 5
        bookings = [self._create_booking(first_name=f"Joe {n}") for n in range(0, COUNT)]
        for b in bookings:
            b.add_to_queue(by_user=b.account)
            assert b.state == BookingState.INFO_COMPLETE

        camp = self.camp
        self.officer_login(officers_factories.create_booking_secretary())
        self.get_url("cciw-officers-booking_queue", camp_id=camp.url_id)

        self.submit('[name="allocate"]')

        for b in bookings:
            b.refresh_from_db()
            assert b.state == BookingState.BOOKED

        self.assertTextPresent("5 places have been allocated, and 5 accounts have been emailed")

        # More detailed tests for `allocate_places_and_notify` below.


class TestBookingQueuePageSL(BookingQueuePageBase, SeleniumBase):
    # This is Selenium only as it requires htmx
    def test_edit_queue_entry(self):
        booking = self._create_booking()
        booking.add_to_queue(by_user=booking.account)
        assert not booking.queue_entry.officer_child
        self.officer_login(user := officers_factories.create_booking_secretary())
        self.get_url("cciw-officers-booking_queue", camp_id=booking.camp.url_id)
        self.click(f'[data-booking-id="{booking.id}"] input[value="Edit queue details"]')
        self.wait_for_ajax()
        self.fill({f'[data-booking-id="{booking.id}"] #id_officer_child': True})
        self.click(f'[data-booking-id="{booking.id}"] input[value="Save"]')
        self.wait_for_ajax()
        booking.refresh_from_db()
        assert booking.queue_entry.officer_child

        action_logs = list(booking.queue_entry.action_logs.all())
        assert len(action_logs) == 2
        log = action_logs[1]
        assert log.staff_user == user
        assert log.action_type == QueueEntryActionLogType.FIELDS_CHANGED
        assert log.details == {
            "fields_changed": [
                {
                    "name": "officer_child",
                    "old_value": False,
                    "new_value": True,
                }
            ]
        }


class TestBookingQueuePageWT(BookingQueuePageBase, WebTestBase):
    pass


@pytest.mark.django_db
def test_booking_queue_track_changes():
    booking = factories.create_booking()
    booking.add_to_queue(by_user=booking.account)
    user = officers_factories.create_booking_secretary()
    queue_entry: BookingQueueEntry = booking.queue_entry
    with queue_entry.track_changes(by_user=user):
        queue_entry.is_active = False
        queue_entry.save()

    logs = list(queue_entry.action_logs.all())
    assert len(logs) == 2
    log = logs[1]
    assert log.details == {
        "fields_changed": [
            {
                "name": "is_active",
                "old_value": True,
                "new_value": False,
            }
        ]
    }
    assert log.staff_user == user


@pytest.mark.django_db
def test_year_config_fetcher(django_assert_num_queries):
    factories.create_year_config(year=2025)
    factories.create_year_config(year=2026)
    fetcher = YearConfigFetcher()

    with django_assert_num_queries(num=1):
        cf = fetcher.lookup_year(2025)
        assert isinstance(cf, YearConfig)

    with django_assert_num_queries(num=0):
        cf2 = fetcher.lookup_year(2025)
        assert isinstance(cf2, YearConfig)
        assert cf2 is cf

    with django_assert_num_queries(num=1):
        assert fetcher.lookup_year(2026) is not None

    with django_assert_num_queries(num=1):
        assert fetcher.lookup_year(2024) is None

    with django_assert_num_queries(num=0):
        assert fetcher.lookup_year(2024) is None


@pytest.mark.django_db
def test_allocate_places(mailoutbox):
    year_config = create_year_config_for_queue_tests()
    camp: Camp = camps_factories.create_camp(
        year=year_config.year, max_campers=5, max_male_campers=5, max_female_campers=5
    )
    booking_sec = officers_factories.create_booking_secretary()

    # Enough accounts and bookings to test all the notification logic.

    accounts = [factories.create_booking_account(name=f"Booker {n}") for n in range(0, 4)]
    bookings = []
    for account in accounts:
        # 2 bookings for each account. With the odd number of places, this
        # leaves one account with one booking confirmed, and one not.
        bookings.extend(
            [
                factories.create_booking(
                    camp=camp, account=account, first_name=f"Joe {n}", last_name=f"Family {account.name}"
                )
                for n in range(0, 2)
            ]
        )

    for idx, booking in enumerate(bookings):
        # We freeze time to after the initial booking period, to give determinism
        # in ranking based on the time
        start = year_config.bookings_close_for_initial_period_on + timedelta(days=1)
        start_dt = datetime(start.year, start.month, start.day)
        with freeze_time(start_dt + timedelta(hours=1 + idx)):
            booking.add_to_queue(by_user=booking.account)

    ranking_result = get_camp_booking_queue_ranking_result(camp=camp, year_config=year_config)
    result = allocate_places_and_notify(ranking_result.bookings, by_user=booking_sec)

    # First 2 accounts get both places accepted,
    # next account gets 1 booking accepted, one declined,
    # Last account gets both declined
    assert result.accepted_account_count == 3
    assert result.accepted_booking_count == 5
    assert result.declined_and_notified_account_count == 2

    assert (
        outbox_count_1 := len(mailoutbox)
    ) == result.accepted_account_count + result.declined_and_notified_account_count

    for booking in bookings:
        booking.refresh_from_db()
        if booking in result.accepted_bookings:
            assert booking.state == BookingState.BOOKED
            assert booking.booking_expires_at is None
            assert booking.queue_entry.declined_notification_sent_at is None
            assert booking.queue_entry.accepted_notification_sent_at is not None
            assert booking.queue_entry.action_logs.filter(action_type=QueueEntryActionLogType.ALLOCATED).exists()
        else:
            assert booking.state == BookingState.INFO_COMPLETE
            assert booking.queue_entry.declined_notification_sent_at is not None
            assert booking.queue_entry.accepted_notification_sent_at is None
            assert booking.queue_entry.action_logs.filter(action_type=QueueEntryActionLogType.DECLINED).exists()
            assert booking.queue_entry.waiting_list_mode

    # Second time: do the same thing.
    ranking_result2 = get_camp_booking_queue_ranking_result(camp=camp, year_config=year_config)
    result2 = allocate_places_and_notify(ranking_result2.bookings, by_user=booking_sec)
    # This time:
    # - No places are allocated, as the camp is full.
    # - No new emails should be sent, because
    #   we notified all the "decline" bookings the first time.
    assert result2.accepted_account_count == 0
    assert result2.accepted_booking_count == 0
    assert result2.declined_and_notified_account_count == 0

    assert len(mailoutbox) == outbox_count_1


@pytest.mark.parametrize("action", ["accept", "cancel", "ignore"])
@pytest.mark.django_db
def test_allocate_places_for_waiting_list(mailoutbox, client: Client, action: Literal["accept", "cancel", "ignore"]):
    year_config = create_year_config_for_queue_tests()
    camp: Camp = camps_factories.create_camp(
        year=year_config.year, max_campers=5, max_male_campers=10, max_female_campers=10
    )
    booking_sec = officers_factories.create_booking_secretary()

    # Enough bookings to fill the camp up:
    initial_bookings = [factories.create_booking() for n in range(0, 5)]
    allocate_bookings_now(initial_bookings)

    # New one, for waiting list.
    booking = factories.create_booking()
    queue_entry = booking.add_to_queue(by_user=booking.account)
    assert queue_entry.waiting_list_from_start
    assert queue_entry.waiting_list_mode

    # A place comes up:
    camp.max_campers = 6
    camp.save()

    # We allocate it:
    ranking_result = get_camp_booking_queue_ranking_result(camp=camp, year_config=year_config)
    result = allocate_places_and_notify(ranking_result.bookings, by_user=booking_sec)

    assert result.accepted_booking_count == 1
    assert result.accepted_bookings == [booking]

    # It is booked:
    booking.refresh_from_db()
    assert booking.state == BookingState.BOOKED

    # But set to expire:
    assert booking.will_expire
    assert booking.booking_expires_at is not None

    # They get emailed:
    assert len(mailoutbox) == 1
    email = mailoutbox[0]

    # The email gives a link to cancel and one to accept:
    accept_url, accept_path, accept_querydata = read_email_url(email, r"https://.*/accept/.*")
    cancel_url, cancel_path, cancel_querydata = read_email_url(email, r"https://.*/cancel/.*")

    # TODO use django client
    if action == "accept":
        response = client.get(accept_path, accept_querydata, follow=True)
        assert response.status_code == 200
        assert b"The place has been confirmed" in response.content

        booking.refresh_from_db()
        assert booking.is_booked
        assert not booking.will_expire

    elif action == "cancel":
        response = client.get(cancel_path, cancel_querydata, follow=True)
        assert response.status_code == 200
        assert b"The place has been cancelled" in response.content

        booking.refresh_from_db()
        assert not booking.is_booked
        assert booking.state == BookingState.CANCELLED_FULL_REFUND
        assert not booking.is_in_queue
        assert booking.shelved

    elif action == "ignore":
        with freeze_time(timezone.now() + settings.BOOKING_EXPIRES_FOR_UNCONFIRMED_BOOKING_AFTER + timedelta(hours=1)):
            expire_bookings()

            booking.refresh_from_db()
            assert not booking.is_booked
            assert booking.state == BookingState.CANCELLED_FULL_REFUND
            assert not booking.is_in_queue

    else:
        assert_never(action)


@pytest.mark.django_db
def test_booking_same_person_on_multiple_camps():
    year_config = create_year_config_for_queue_tests()
    year: int = year_config.year
    camp_1: Camp = camps_factories.create_camp(year=year)
    camp_2: Camp = camps_factories.create_camp(year=year)

    # Enough accounts and bookings to test all the notification logic.

    account = factories.create_booking_account()

    booking_1 = factories.create_booking(camp=camp_1, account=account, first_name="Joe", last_name="Bloggs")
    # Note the lower case below, we should still get a match
    booking_2 = factories.create_booking(camp=camp_2, account=account, first_name="joe", last_name="bloggs")

    problems1 = get_booking_problems(booking_2)
    messages1 = [p.description for p in problems1]
    msg = 'You are trying to book places for "joe bloggs" on more than one camp.'
    assert len([True for m in messages1 if msg in m]) == 1

    # If booking_1 is put on shelf, there is no problem.
    booking_1.move_to_shelf()

    problems2 = get_booking_problems(booking_2)
    assert len(problems2) == len(problems1) - 1
    messages2 = [p.description for p in problems2]
    assert len([True for m in messages2 if msg in m]) == 0


@pytest.mark.django_db
def test_booking_fuzzy_camper_id_strict():
    booking_1 = factories.create_booking()
    assert booking_1.fuzzy_camper_id_strict_unsaved == booking_1.fuzzy_camper_id_strict

    booking_2 = Booking(
        account=booking_1.account,
        first_name=booking_1.first_name,
        last_name=booking_1.last_name,
        birth_date=booking_1.birth_date,
    )
    assert booking_2.fuzzy_camper_id_strict_unsaved == booking_1.fuzzy_camper_id_strict


@pytest.mark.django_db
def test_add_to_queue_after_camp_full():
    # Bookings that are added to a queue after a camp is full
    # are marked as 'waiting_list_from_start'
    camp: Camp = camps_factories.create_camp(max_campers=5)
    initial_bookings = [factories.create_booking(camp=camp) for _ in range(0, 5)]
    for initial_booking in initial_bookings:
        initial_queue_entry = initial_booking.add_to_queue(by_user=initial_booking.account)
        assert not initial_queue_entry.waiting_list_from_start
        assert not initial_queue_entry.waiting_list_mode
        assert initial_queue_entry.will_send_declined_notification

    # Make camp full:
    allocate_bookings_now(initial_bookings)

    # Another one faces a camp already full
    booking = factories.create_booking(camp=camp)
    queue_entry = booking.add_to_queue(by_user=booking.account)
    assert queue_entry.waiting_list_from_start
    assert queue_entry.waiting_list_mode
    assert not queue_entry.will_send_declined_notification

    # If we cancel a place, then add again, and it faces an already full camp,
    # then waiting_list_from_start should be adjusted.
    camp.max_campers = 3
    camp.save()

    b1 = initial_bookings[0]
    b1.withdraw_from_queue(by_user=b1.account)

    q1 = b1.add_to_queue(by_user=b1.account)
    assert q1.waiting_list_from_start

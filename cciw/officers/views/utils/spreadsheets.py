import openpyxl

from cciw.officers.views.utils.data_retention import (
    DATA_RETENTION_NOTICES_TXT,
    DataRelation,
    DataRetentionRule,
    SensitiveDownloadResponse,
)
from cciw.utils import xl
from cciw.utils.spreadsheet import ExcelBuilder


def spreadsheet_response(
    builder: ExcelBuilder,
    filename: str,
    *,
    rule: DataRetentionRule | None,
    data_relation: DataRelation,
) -> SensitiveDownloadResponse:
    output = builder.to_bytes()

    if rule is not None:
        workbook: openpyxl.Workbook = xl.workbook_from_bytes(builder.to_bytes())
        sheet = workbook.create_sheet("Notice", 0)
        c_header = sheet.cell(1, 1)
        c_header.value = "Data retention notice:"
        c_header.font = xl.header_font

        for row_idx, line in enumerate(notice_to_lines(rule), start=3):
            c = sheet.cell(row_idx, 1)
            c.value = line
            c.font = xl.default_font
        sheet.column_dimensions["A"].width = 100

        output = xl.workbook_to_bytes(workbook)
    # All spreadsheets are assumed to be sensitive by default,
    # NoSensitiveData can be used for those that aren't.
    return SensitiveDownloadResponse(
        output, content_type=builder.mimetype, data_relation=data_relation, filename=f"{filename}.{builder.file_ext}"
    )


def notice_to_lines(rule: DataRetentionRule, /) -> list[str]:
    txt = DATA_RETENTION_NOTICES_TXT[rule]
    return list(txt.split("\n"))

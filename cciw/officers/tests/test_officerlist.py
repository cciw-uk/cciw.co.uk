import io
from datetime import date, timedelta

import openpyxl
from django.conf import settings
from django.core import mail
from openpyxl.utils import get_column_letter

from cciw.accounts.models import User
from cciw.cciwmain.tests import factories as camp_factories
from cciw.cciwmain.tests.base import SiteSetupMixin
from cciw.officers.create import create_officer
from cciw.officers.models import Application
from cciw.officers.tests import factories
from cciw.officers.utils import camp_serious_slacker_list, officer_data_to_spreadsheet
from cciw.utils.spreadsheet import ExcelSimpleBuilder
from cciw.utils.tests.base import TestBase
from cciw.utils.tests.webtest import SeleniumBase, WebTestBase


class TestCreate(TestBase):
    def test_create(self):
        user = create_officer("Joe", "Bloggs", "joebloggs@example.com")

        user = User.objects.get(id=user.id)
        assert user.is_staff
        assert len(mail.outbox) == 1
        assert user.last_login is None


class TestExport(TestBase):
    def test_export_no_application(self):
        """
        Test that the export data view generates an Excel file with all the data
        we expect if there is no application form.
        """
        camp = camp_factories.create_camp(
            officers=[officer := factories.create_officer()], officers_role="Tent Officer"
        )
        first_names = [o.first_name for o in [officer]]

        assert Application.objects.all().count() == 0

        workbook = officer_data_to_spreadsheet(camp, ExcelSimpleBuilder()).to_bytes()

        assert workbook is not None
        wkbk: openpyxl.Workbook = openpyxl.load_workbook(io.BytesIO(workbook))
        wksh = wkbk.worksheets[0]

        # Spot checks on different types of data
        # From User model
        assert wksh.cell(1, 1).value == "First name"
        assert wksh.cell(2, 1).value in first_names

        # From Invitation model
        assert wksh.cell(1, 4).value == "Role"
        assert wksh.cell(2, 4).value == "Tent Officer"

    def test_export_with_application(self):
        """
        Test that the export data view generates an Excel file with all the data
        we expect if there are application forms.
        """
        camp = camp_factories.create_camp(officers=[officer := factories.create_officer()])
        factories.create_application(year=camp.year, officer=officer, address_firstline="123 The Way")

        workbook = officer_data_to_spreadsheet(camp, ExcelSimpleBuilder()).to_bytes()

        wkbk: openpyxl.Workbook = openpyxl.load_workbook(io.BytesIO(workbook))
        wksh = wkbk.worksheets[0]

        # Check data from Application model
        assert wksh.cell(1, 5).value == "Address"
        assert "123 The Way" in [c.value for c in wksh[get_column_letter(5)]]


class TestSlackers(TestBase):
    def test_serious_slackers(self):
        officer1 = factories.create_officer()
        officer2 = factories.create_officer()
        camp1 = camp_factories.create_camp(year=date.today().year - 2, officers=[officer1, officer2])
        camp2 = camp_factories.create_camp(year=date.today().year - 1, officers=[officer1, officer2])

        # Officer 1 submitted an Application, but officer 2 did not
        app = officer1.applications.create(
            date_saved=camp1.start_date - timedelta(days=10),
            finished=True,
        )

        # Officer 1 submitted references, but officer 2 did not
        factories.create_complete_reference(app.referees[0])
        factories.create_complete_reference(app.referees[1])

        # Officer 1 got a DBS done, but officer 2 did not
        officer1.dbs_checks.create(
            dbs_number="123456",
            completed=camp1.start_date - timedelta(days=5),
        )

        serious_slackers = camp_serious_slacker_list(camp2)

        assert serious_slackers == [
            {
                "officer": officer2,
                "missing_application_forms": [camp1],
                "missing_references": [camp1],
                "missing_dbss": [camp1],
                "last_good_apps_year": None,
                "last_good_refs_year": None,
                "last_good_dbss_year": None,
            }
        ]


class TestApplicationFormStatusPAge(SiteSetupMixin, WebTestBase):
    def test_page(self):
        camp = camp_factories.create_camp(
            leader=(leader := factories.create_officer()),
            officers=[(officer := factories.create_officer())],
        )
        self.officer_login(leader)
        self.get_url("cciw-officers-officer_application_status", camp_id=camp.url_id)
        self.assertTextPresent(officer.full_name, within="#id_noapplicationform")


class CampRoleSetupMixin:
    def setUp(self):
        super().setUp()
        self.camp_role1 = factories.create_camp_role(name="Tent Officer")
        self.camp_role2 = factories.create_camp_role(name="Assistant Leader")


class TestOfficerListPage(SiteSetupMixin, CampRoleSetupMixin, SeleniumBase):
    def remove_button_selector(self, officer):
        return f'[data-officer-id="{officer.id}"] [name="remove"]'

    def edit_button_selector(self, officer):
        return f'[data-officer-id="{officer.id}"] [name="edit"]'

    def save_button_selector(self, officer):
        return f'[data-officer-id="{officer.id}"] [name="save"]'

    def cancel_button_selector(self, officer):
        return f'[data-officer-id="{officer.id}"] [name="cancel"]'

    def resend_email_button_selector(self, officer):
        return f'[data-officer-id="{officer.id}"] [name="resend_email"]'

    def test_add_with_new_role(self):
        camp = camp_factories.create_camp(
            leader=(leader := factories.create_officer()),
        )
        officer = factories.create_officer()
        self.officer_login(leader)
        self.get_url("cciw-officers-officer_list", camp_id=camp.url_id)

        # Check initial:
        assert officer not in camp.officers.all()
        self.assertTextPresent(officer.email)
        self.assertTextPresent("No officers added yet")

        # Action:
        self.fill(
            {
                f"#id_chooseofficer_{ officer.id }": True,
                "#id_chooseofficer-role": str(self.camp_role1.id),
            }
        )
        self.click("input[name=add_new_role]")

        # UI check (first, because assertTextPresent does a helpful wait for us)
        self.assertTextPresent(self.camp_role1.name, within=f'[data-officer-id="{officer.id}"]')

        # DB check:
        assert officer in camp.officers.all()
        assert camp.invitations.get(officer=officer).role == self.camp_role1

    def test_add_with_previous_role(self):
        camp = camp_factories.create_camp(
            leader=(leader := factories.create_officer()),
        )
        officer = factories.create_officer()
        previous_camp = camp_factories.create_camp(camp_name=camp.camp_name, leader=leader, year=camp.year - 1)
        factories.add_officers_to_camp(
            previous_camp, [officer], role=factories.get_or_create_camp_role(name="Kitchen helper")
        )
        assert not camp.invitations.all()

        self.officer_login(leader)
        self.get_url("cciw-officers-officer_list", camp_id=camp.url_id)

        # Check initial:
        self.assertTextPresent("No officers added yet")

        # Action:
        self.fill({f"#id_chooseofficer_{ officer.id }": True})
        self.click("input[name=add_previous_role]")

        # UI check:
        self.assertTextPresent("Kitchen helper", within=f'[data-officer-id="{officer.id}"]')

        # DB check:
        assert officer in camp.officers.all()
        assert camp.invitations.get(officer=officer).role.name == "Kitchen helper"

    def test_remove(self):
        camp = camp_factories.create_camp(
            leader=(leader := factories.create_officer()),
            officers=[officer := factories.create_officer(first_name="Zog", last_name="The Dragon")],
        )

        self.officer_login(leader)
        self.get_url("cciw-officers-officer_list", camp_id=camp.url_id)

        # Check initial:
        assert officer in camp.officers.all()
        self.assertTextAbsent(officer.full_name, within=".chooseofficers")
        self.assertTextPresent(officer.full_name, within=f'[data-officer-id="{officer.id}"]')
        self.assertTextPresent(officer.email, within=f'[data-officer-id="{officer.id}"]')

        # Action:
        self.click(self.remove_button_selector(officer))

        # Assert
        self.assertTextPresent(officer.full_name, within=".chooseofficers")
        assert not self.is_element_present(f'[data-officer-id="{officer.id}"]')

        assert officer not in camp.officers.all()

    def test_resend_email(self):
        camp = camp_factories.create_camp(
            leader=(leader := factories.create_officer()),
            officers=[officer := factories.create_officer()],
        )

        self.officer_login(leader)
        self.get_url("cciw-officers-officer_list", camp_id=camp.url_id)

        # Action:
        self.click(self.resend_email_button_selector(officer), expect_alert=True)
        self.accept_alert()
        self.wait_until(lambda *args: len(mail.outbox) > 0)
        self.assertTextPresent("Sent!", within=f'[data-officer-id="{officer.id}"]')
        (m,) = mail.outbox
        assert officer.first_name in m.body
        assert "https://" + settings.PRODUCTION_DOMAIN + "/officers/" in m.body

    def test_edit(self):
        camp = camp_factories.create_camp(
            leader=(leader := factories.create_officer()),
        )
        officer = factories.create_officer()

        (invitation,) = factories.add_officers_to_camp(camp, [officer], role=self.camp_role1)

        self.officer_login(leader)
        self.get_url("cciw-officers-officer_list", camp_id=camp.url_id)

        # Change officer details, not role (checking that role is propagated)
        self.click(self.edit_button_selector(officer))
        self.fill(
            {
                "#id_first_name": "Altered",
                "#id_last_name": "Name",
                "#id_email": "alteredemail@somewhere.com",
            }
        )
        self.click(self.save_button_selector(officer))

        self.wait_until_loaded(self.edit_button_selector(officer))

        # Test DB
        officer.refresh_from_db()
        assert officer.first_name == "Altered"
        assert officer.last_name == "Name"
        assert officer.email == "alteredemail@somewhere.com"
        invitation.refresh_from_db()
        assert invitation.role == self.camp_role1  # unchanged

        # Test UI:
        self.assertTextPresent("alteredemail@somewhere.com")

        # Change role, not user details (checking user details are propagated)
        self.click(self.edit_button_selector(officer))
        self.fill({"#id_role": str(self.camp_role2.id)})
        self.click(self.save_button_selector(officer))
        self.wait_until_loaded(self.edit_button_selector(officer))

        officer.refresh_from_db()
        assert officer.email == "alteredemail@somewhere.com"
        invitation.refresh_from_db()
        assert invitation.role == self.camp_role2

    def test_edit_validation(self):
        camp = camp_factories.create_camp(
            leader=(leader := factories.create_officer()),
            officers=[officer := factories.create_officer()],
        )

        self.officer_login(leader)
        self.get_url("cciw-officers-officer_list", camp_id=camp.url_id)

        self.click(self.edit_button_selector(officer))
        self.fill({"#id_email": "bademail"})
        self.click(self.save_button_selector(officer))

        # Test DB
        officer = User.objects.get(id=officer.id)
        assert officer.email != "bademail"

        # Test UI:
        assert self.is_element_displayed(self.save_button_selector(officer))

    def test_edit_cancel(self):
        camp = camp_factories.create_camp(
            leader=(leader := factories.create_officer()),
            officers=[(officer := factories.create_officer())],
        )

        self.officer_login(leader)
        self.get_url("cciw-officers-officer_list", camp_id=camp.url_id)

        self.click(self.edit_button_selector(officer))
        self.fill(
            {
                "#id_first_name": "Altered",
                "#id_last_name": "Name",
            }
        )
        self.click(self.cancel_button_selector(officer))

        self.wait_until_loaded(self.edit_button_selector(officer))

        # Test DB
        officer.refresh_from_db()
        assert officer.first_name != "Altered"
        assert officer.last_name != "Name"

    def test_new_officer_roundtrip(self):
        camp = camp_factories.create_camp(
            leader=(leader := factories.create_officer()),
        )
        self.officer_login(leader)
        self.get_url("cciw-officers-officer_list", camp_id=camp.url_id)
        self.click(".newofficers summary")
        self.follow_link(".newofficers a")

        self.fill(
            {
                "#id_first_name": "Mary",
                "#id_last_name": "Andrews",
                "#id_email": "mary@andrews.com",
            }
        )
        self.submit("input[type=submit]")
        self.assertTextPresent("Officer Mary Andrews has been added to the system")
        self.assertTextPresent("Don't forget to choose a role and add them")
        officer = User.objects.get(email="mary@andrews.com")

        # Officer should be selected already, just need to choose a role
        self.fill({"#id_chooseofficer-role": str(self.camp_role1.id)})
        self.click("input[name=add_new_role]")

        # UI check:
        self.assertTextPresent("Mary Andrews", within=f'[data-officer-id="{officer.id}"]')

        # DB check:
        assert officer in camp.officers.all()
        assert camp.invitations.get(officer=officer).role == self.camp_role1


class TestNewOfficerPage(SiteSetupMixin, WebTestBase):
    CONFIRM_BUTTON = "input[name=confirm]"

    def setUp(self):
        super().setUp()
        mail.outbox = []

    def get_page(self):
        self.get_url("cciw-officers-create_officer")

    def test_permissions(self):
        camp_factories.create_camp(leader=(leader := factories.create_officer()), future=True)
        self.get_page()
        assert self.is_element_present("body.login")
        self.officer_login(leader)
        self.get_page()
        assert not self.is_element_present("body.login")
        self.assertTextPresent("Enter details for officer")

    def _access_officer_list_page(self):
        camp_factories.create_camp(leader=(leader := factories.create_officer()), future=True)
        self.officer_login(leader)
        self.get_page()

    def test_success(self):
        self._access_officer_list_page()
        self.fill(
            {
                "#id_first_name": "Mary",
                "#id_last_name": "Andrews",
                "#id_email": "mary@andrews.com",
            }
        )
        self.submit("input[type=submit]")
        self._assert_created()

    def test_duplicate_user(self):
        factories.create_officer(
            username="maryandrews", first_name="Mary", last_name="Andrews", email="mary@andrews.com"
        )
        self._access_officer_list_page()
        self.fill(
            {
                "#id_first_name": "Mary",
                "#id_last_name": "Andrews",
                "#id_email": "mary@andrews.com",
            }
        )
        self.submit("input[type=submit]")
        self.assertTextPresent("A user with that name and email address already exists")
        assert not self.is_element_present(self.CONFIRM_BUTTON)

    def test_duplicate_name(self):
        factories.create_officer(
            username="maryandrews", first_name="Mary", last_name="Andrews", email="mary@otheremail.com"
        )
        self._access_officer_list_page()
        self.fill(
            {
                "#id_first_name": "Mary",
                "#id_last_name": "Andrews",
                "#id_email": "mary@andrews.com",
            }
        )
        self.submit("input[type=submit]")
        self.assertTextPresent("A user with that first name and last name already exists")
        self.submit(self.CONFIRM_BUTTON)
        self._assert_created()

    def test_duplicate_email(self):
        factories.create_officer(
            username="mikeandrews", first_name="Mike", last_name="Andrews", email="mary@andrews.com"
        )
        self._access_officer_list_page()
        self.fill(
            {
                "#id_first_name": "Mary",
                "#id_last_name": "Andrews",
                "#id_email": "mary@andrews.com",
            }
        )
        self.submit("input[type=submit]")
        self.assertTextPresent("A user with that email address already exists")
        self.submit(self.CONFIRM_BUTTON)
        self._assert_created()

    def _assert_created(self):
        u = User.objects.get(email="mary@andrews.com", first_name="Mary")
        assert u.first_name == "Mary"
        assert u.last_name == "Andrews"
        c = User.objects.filter(first_name="Mary", last_name="Andrews").count()
        username = "maryandrews" + (str(c) if c > 1 else "")

        assert u.username == username
        assert len(mail.outbox) == 1
        m = mail.outbox[0]
        assert "Hi Mary" in m.body
        assert "https://" + settings.PRODUCTION_DOMAIN + "/officers/" in m.body
